import datetime
import itertools
import operator

import openpyxl
import openpyxl.cell
import openpyxl.utils

from django.contrib.auth.models import User
from django.core.exceptions import ValidationError as DjangoValidationError
from django.db import (
    IntegrityError,
    transaction,
)

from django.http import (
    HttpResponse,
    JsonResponse,
)
from django.utils import timezone
from rest_framework import status

from rest_framework.exceptions import ValidationError
from rest_framework.fields import (
    BooleanField,
    CharField,
    ChoiceField,
    DateField,
    DecimalField,
    FileField,
    ImageField,
    IntegerField,
    ListField,
    TimeField,
)
from rest_framework.serializers import (
    Serializer,
    SerializerMetaclass,
)

from doc_templates.http_responses import xlsx_content_response

from the_redhuman_is.async_utils.telephony import (
    get_call_history,
    start_call,
)
from the_redhuman_is.auth import user_in_group
from the_redhuman_is.models.delivery import (
    DeliveryRequest,
    ItemWorkerRejection,
    WorkerZone,
    ZoneGroup,
)
from the_redhuman_is.models.worker import (
    BannedWorker,
    Worker,
    WorkerRating,
)
from the_redhuman_is.services.delivery import (
    actions,
    retrieve,
)
from the_redhuman_is.services.delivery.utils import (
    DeliveryWorkflowError,
    HoursError,
    ObjectNotFoundError,
    normalize_driver_name,
    normalize_driver_phones,
)
from the_redhuman_is.views.backoffice_app.auth import bo_api
from the_redhuman_is.views.utils import (
    _get_value,
    get_first_last_day,
)
from utils import (
    date_time,
    phone as phone_utils,
)
from utils.functools import merge_dicts


class RequestQuerySerializer(Serializer):
    request = IntegerField(min_value=1, source='request_id')


@bo_api(['POST'])
def call_to_driver(request):
    serializer = RequestQuerySerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    try:
        driver_name, driver_phones = DeliveryRequest.objects.values_list(
            'driver_name',
            'driver_phones',
        ).get(
            pk=serializer.validated_data['request_id']
        )
    except DeliveryRequest.DoesNotExist as e:
        raise ObjectNotFoundError from e

    phones = phone_utils.extract_phones(driver_phones)
    if not phones:
        raise DeliveryWorkflowError('В заявке не указаны телефоны водителя.')

    start_call(request.user.username, driver_name, phones)

    return HttpResponse(status=status.HTTP_202_ACCEPTED)


class WorkerQuerySerializer(Serializer):
    worker = IntegerField(min_value=1, source='worker_id')


@bo_api(['POST'])
def call_to_worker(request):
    serializer = WorkerQuerySerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    worker_id = serializer.validated_data['worker_id']
    try:
        tel_number, full_name = Worker.objects.with_full_name(
        ).values_list(
            'tel_number',
            'full_name',
        ).get(
            pk=worker_id
        )
    except Worker.DoesNotExist as e:
        raise ObjectNotFoundError from e

    phones = phone_utils.extract_phones(tel_number)

    start_call(request.user.username, full_name, phones)
    WorkerRating.objects.update_or_create(
        worker_id=worker_id,
        defaults={
            'last_call': timezone.now()
        }
    )

    return HttpResponse(status=status.HTTP_202_ACCEPTED)


CALL_TYPE_DESCRIPTION = {
    'out_click_to_call': 'по кнопке',
    'out_manual': 'набор вручную',
    'inc': 'входящий',
    'unknown': 'неизвестно'
}

CALL_DIRECTION = {
    'out': 'исходящий',
    'inc': 'входящий',
}
CONTACT_TYPE = {
    'D': 'В',
    'W': 'Р',
}


def worker_driver_calls_history(first_day, last_day):
    delivery_requests = DeliveryRequest.objects.filter(
        date__range=(first_day, last_day)
    ).with_zone_name(
    ).order_by(
        'date',
        'pk',
    ).values(
        'date',
        'driver_name',
        'driver_phones',
        'zone_name',
    )
    driver_phones = {}
    for day, day_requests in itertools.groupby(
            delivery_requests,
            key=operator.itemgetter('date')
    ):
        day_phones = {}
        for request in day_requests:
            zone_name = request['zone_name']
            for phone in phone_utils.extract_phones(request['driver_phones']):
                if phone not in day_phones or zone_name is not None:
                    day_phones[phone] = (request['driver_name'], zone_name)
        driver_phones[day] = day_phones

    workers = Worker.objects.all(
    ).with_full_name(
    ).order_by(
        'pk'
    ).values(
        'tel_number',
        'full_name',
        'workerzone__zone__name',
    )
    worker_phones = {
        worker['tel_number']: (worker['full_name'], worker['workerzone__zone__name'])
        for worker in workers
    }

    users = {}

    def _user_first_name(username):
        if username not in users:
            try:
                user = User.objects.get(username=username)
                users[username] = user.first_name
            except User.DoesNotExist:
                users[username] = username
        return users[username]

    history = get_call_history(first_day, last_day)

    for item in history:
        item['dispatchers'] = list(dict.fromkeys(item['users']))

        day = item['start_time'].date()
        phones = item['phones']
        try:
            try:
                contact = driver_phones[day][phones[0]]
                item['contact_type'] = 'D'
            except KeyError:
                contact = worker_phones[phones[0]]
                item['contact_type'] = 'W'
        except (KeyError, IndexError):
            contact = None, None
            item['contact_type'] = None
        item['contact_name'], item['zone'] = contact

        item['dispatcher_names'] = [
            _user_first_name(username) for username in item['dispatchers']
        ]

    return history


def calls_report(request):
    first_day, last_day = get_first_last_day(request)
    min_day = datetime.date(day=18, month=7, year=2021)
    first_day = max(min_day, first_day)
    last_day = max(min_day, last_day)

    history = worker_driver_calls_history(first_day, last_day)

    captions = (
        ('Дата', 12),
        ('Время', 9),
        ('Сотрудник (ФИО)', 18),
        ('Тип вызова', 16),
        ('Тип вызов (вх/исх)', 12),
        ('Контактное лицо (фио)', 40),
        ('Город', 17),
        ('Контакт', 7),
        ('Номер (контактное лицо)', 15),
        ('Статус вызова (дозв, не дозв, не соединился и т. д.)', 10),
        ('Длительность соединения', 8),
        ('Время разговора', 8),
        ('Запись звонка', 13),
    )

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet()

    for col, (caption, width) in enumerate(captions, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    ws.append([openpyxl.cell.WriteOnlyCell(ws, value=caption) for caption, _ in captions])

    def _make_cell(value, number_format=None, link=None):
        cell = openpyxl.cell.WriteOnlyCell(ws, value=value)
        if number_format is not None:
            cell.number_format = number_format
        if link is not None:
            cell.hyperlink = link
        return cell

    def _make_cells(row):
        return [
            _make_cell(value, number_format=cell_format.get(j))
            for j, value in enumerate(row, start=1)
        ]

    cell_format = {
        1: 'DD.MM.YYYY',
        2: 'HH:MM:SS',
        9: '@',
    }

    for item in history:
        ws.append(
            _make_cells([
                item['start_time'].date(),
                item['start_time'].time(),
                ', '.join(item['dispatcher_names']),
                CALL_TYPE_DESCRIPTION.get(item['call_type'], 'непонятно'),
                CALL_DIRECTION.get(item['call_type'][:3], 'неизвестно'),
                item['contact_name'] or '-',
                item['zone'] or '-',
                CONTACT_TYPE.get(item['contact_type'], '-'),
                ', '.join(item['phones']),
                'дозв' if item['successful'] else 'не дозв',
                item['total_duration'],
                item['duration'],
            ]) + [
                _make_cell(value='запись', link=link)
                for link in item['record_urls']
            ]
        )

    return xlsx_content_response(
        wb,
        f'Звонки с {date_time.string_from_date(first_day)}'
        f' по {date_time.string_from_date(last_day)}.xlsx'
    )


@bo_api(['POST'])
@user_in_group('Бан рабочих')
def ban_worker(request):
    try:
        worker = Worker.objects.get(pk=int(_get_value(request, 'worker')))
        BannedWorker.objects.create(worker=worker)
        return JsonResponse({
            'status_code': Worker.BANNED
        })
    except (KeyError, ValueError, Worker.DoesNotExist, IntegrityError):
        return JsonResponse({}, status=400)


@bo_api(['POST'])
@user_in_group('Бан рабочих')
def unban_worker(request):
    try:
        worker = Worker.objects.select_related(
            'banned'
        ).get(
            pk=int(_get_value(request, 'worker'))
        )
        worker.banned.delete()

        status_code = Worker.objects.with_turnout_status_code(
        ).values_list(
            'status_code',
            flat=True
        ).get(
            pk=worker.pk
        )
        return JsonResponse({
            'status_code': status_code
        })
    except (KeyError, ValueError, Worker.DoesNotExist, BannedWorker.DoesNotExist):
        return JsonResponse({}, status=400)


@bo_api(['POST'])
def set_worker_zone(request, pk):
    try:
        zone = ZoneGroup.objects.get(pk=int(_get_value(request, 'zone')))
        worker = Worker.objects.select_related('workerzone').get(pk=pk)
        with transaction.atomic():
            try:
                worker.workerzone.zone = zone
                worker.workerzone.save(update_fields=['zone'])
            except WorkerZone.DoesNotExist:
                WorkerZone.objects.create(worker=worker, zone=zone)
        return JsonResponse({
            'zone': {
                'id': zone.pk,
                'name': zone.name,
                'code': zone.code,
            }
        })
    except (KeyError, ValueError, Worker.DoesNotExist):
        return JsonResponse({}, status=400)


# new methods

class ForceCommit(metaclass=SerializerMetaclass):
    force_commit = BooleanField(default=False)


@bo_api(['GET'])
def list_requests(request):
    return JsonResponse(
        retrieve.get_delivery_request_list(request.user, request.GET)
    )


class RequestDetailQuerySerializer(Serializer):
    id = IntegerField(source='request_id', min_value=1)
    with_photos = BooleanField(default=False)


@bo_api(['GET'])
def request_detail(request):
    serializer = RequestDetailQuerySerializer(data=request.GET)
    serializer.is_valid(raise_exception=True)
    return JsonResponse(
        {
            'data': retrieve.get_delivery_request_detail(
                user=request.user,
                **serializer.validated_data
            ),
        }
    )


class DeliveryItemCreateSerializer(Serializer):
    interval_begin = TimeField()
    interval_end = TimeField()
    code = CharField()
    mass = DecimalField(min_value=0, max_digits=10, decimal_places=3)
    volume = DecimalField(min_value=0, max_digits=10, decimal_places=3)
    max_size = DecimalField(min_value=0, max_digits=10, decimal_places=3, allow_null=True)
    place_count = IntegerField(min_value=0)
    shipment_type = CharField()
    address = CharField()
    has_elevator = BooleanField(allow_null=True)
    floor = IntegerField(allow_null=True)
    carrying_distance = IntegerField(allow_null=True)
    workers_required = IntegerField(min_value=1, default=1)
    confirmed_timepoint = TimeField(allow_null=True, default=None)


class DeliveryRequestCreateSerializer(Serializer):
    request = IntegerField(min_value=1, source='request_id', default=None)
    customer = IntegerField(min_value=1, source='customer_id')
    location = IntegerField(min_value=1, source='location_id', allow_null=True, default=None)
    date = DateField(input_formats=[date_time.DATE_FORMAT])
    driver_name = CharField(default='', allow_blank=True)
    driver_phones = CharField(default='', allow_blank=True)
    items = DeliveryItemCreateSerializer(many=True, allow_empty=False)
    merge = BooleanField(default=True)
    notify_dispatchers = BooleanField(default=True)

    def validate_driver_name(self, value):
        return normalize_driver_name(value)

    def validate_driver_phones(self, value):
        return normalize_driver_phones(value)


@bo_api(['POST'])
def create_request(request):
    serializer = DeliveryRequestCreateSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    delivery_request = actions.create_delivery_request(
        user=request.user,
        **serializer.validated_data
    )
    return JsonResponse(
        {
            'data': retrieve.get_delivery_request_detail(
                request_id=delivery_request.pk,
                user=request.user,
            ),
            'messages': [],
            'confirmation_required': False,
        }
    )


class DeliveryRequestUpdateSerializer(Serializer, ForceCommit):
    request = IntegerField(min_value=1, source='request_id', default=None)
    customer = IntegerField(min_value=1)
    location = IntegerField(min_value=1)
    date = DateField(input_formats=[date_time.DATE_FORMAT])
    driver_name = CharField()
    driver_phones = CharField()
    customer_confirmation = BooleanField()
    service = IntegerField(allow_null=True)
    comment = CharField(allow_null=True)
    status = CharField()
    operator = IntegerField()
    is_private = BooleanField()

    def validate_driver_name(self, value):
        return normalize_driver_name(value)

    def validate_driver_phones(self, value):
        return normalize_driver_phones(value)

    def validate_comment(self, value):
        if value is None:
            return ''
        return value

    def validate(self, attrs):
        if 'request_id' not in attrs:
            raise ValidationError('Номер заявки является обязательным полем.')
        n_fields = len(attrs.keys() - {'request_id', 'force_commit'})
        if n_fields > 1:
            raise ValidationError('Одновременно можно обновлять только одно поле.')
        elif n_fields == 0:
            raise ValidationError(
                'Отсутствует обновляемое поле. Доступные для редактирования поля: '
                'customer, location, date, driver_name, driver_phones, customer_confirmation, '
                'service, comment, status, operator, is_private.'
            )
        return attrs

    def validate_status(self, value):
        if value not in DeliveryRequest.MANUAL_STATUSES:
            raise ValidationError(
                'Вручную можно установить только следующие статусы: "Новая", "Не принята в'
                'работу", "Отмена", "Перезвонит сам", "Нет ответа" и "Отмена с оплатой".'
            )
        return value


@bo_api(['POST'])
def update_request(request):
    serializer = DeliveryRequestUpdateSerializer(data=request.data, partial=True)
    serializer.is_valid(raise_exception=True)
    request_id = serializer.validated_data.pop('request_id')
    force_commit = serializer.validated_data.pop('force_commit', False)
    field = next(iter(serializer.validated_data.keys()))
    value = next(iter(serializer.validated_data.values()))
    messages, confirmation_required = actions.update_delivery_request(
        request_id,
        field,
        value,
        request.user,
        force_commit
    )
    return JsonResponse(
        {
            'data': retrieve.get_delivery_request_detail(
                request_id=request_id,
                user=request.user,
            ),
            'messages': messages,
            'confirmation_required': confirmation_required,
        }
    )


class DeliveryItemUpdateParamsSerializer(Serializer, ForceCommit):
    request = IntegerField(min_value=1, source='request_id')
    item = IntegerField(min_value=1, source='item_id')


class DeliveryItemUpdateFieldsSerializer(DeliveryItemCreateSerializer):
    workers_required = IntegerField(min_value=0)
    first = BooleanField()

    def validate(self, attrs):
        n_fields = len(attrs.keys() - {'force_commit'})
        if n_fields > 1:
            raise ValidationError('Одновременно можно обновлять только одно поле.')
        elif n_fields == 0:
            raise ValidationError(
                'Отсутствует обновляемое поле. Доступные для редактирования поля: '
                'interval_begin, interval_end, code, mass, volume, max_size, place_count, '
                'shipment_type, address, has_elevator, floor, carrying_distance, '
                'workers_required, confirmed_timepoint.'
            )
        return attrs


@bo_api(['POST'])
def update_item(request):
    ser_params = DeliveryItemUpdateParamsSerializer(data=request.data)
    ser_params.is_valid()
    ser_field = DeliveryItemUpdateFieldsSerializer(data=request.data, partial=True)
    ser_field.is_valid()
    if ser_params.errors or ser_field.errors:
        raise ValidationError(merge_dicts(ser_params.errors, ser_field.errors))
    field = next(iter(ser_field.validated_data.keys()))
    value = next(iter(ser_field.validated_data.values()))
    _, messages, confirmation_required = actions.update_delivery_item(
        field=field,
        value=value,
        user=request.user,
        **ser_params.validated_data
    )
    return JsonResponse(
        {
            'data': retrieve.get_delivery_request_detail(
                request_id=ser_params.validated_data['request_id'],
                user=request.user,
            ),
            'messages': messages,
            'confirmation_required': confirmation_required,
        }
    )


class RequestWorkerSerializer(Serializer):
    request = IntegerField(min_value=1, source='request_id')
    worker = IntegerField(min_value=1, source='worker_id')


@bo_api(['POST'])
def add_request_worker(request):
    serializer = RequestWorkerSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    actions.add_request_worker(
        user=request.user,
        **serializer.validated_data
    )
    return JsonResponse(
        {
            'data': retrieve.get_delivery_request_detail(
                request_id=serializer.validated_data['request_id'],
                user=request.user,
            ),
            'messages': [],
            'confirmation_required': False
        }
    )


@bo_api(['POST'])
def confirm_request_worker(request):
    serializer = RequestWorkerSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    actions.confirm_request_worker(
        user=request.user,
        **serializer.validated_data
    )
    return JsonResponse(
        {
            'data': retrieve.get_delivery_request_detail(
                request_id=serializer.validated_data['request_id'],
                user=request.user,
            ),
            'messages': [],
            'confirmation_required': False,
        }
    )


class RequestWorkerConfirmSerializer(RequestWorkerSerializer, ForceCommit):
    pass


@bo_api(['POST'])
def remove_request_worker(request):
    serializer = RequestWorkerConfirmSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    _, messages, confirmation_required = actions.remove_request_worker(
        user=request.user,
        **serializer.validated_data
    )
    return JsonResponse(
        {
            'data': retrieve.get_delivery_request_detail(
                request_id=serializer.validated_data['request_id'],
                user=request.user,
            ),
            'messages': messages,
            'confirmation_required': confirmation_required,
        }
    )


class DeliveryItemAddSerializer(DeliveryItemCreateSerializer):
    # + fields from Item
    request = IntegerField(min_value=1, source='request_id')


@bo_api(['POST'])
def create_delivery_item(request):
    serializer = DeliveryItemAddSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    actions.create_delivery_item(
        user=request.user,
        **serializer.validated_data
    )
    return JsonResponse(
        {
            'data': retrieve.get_delivery_request_detail(
                request_id=serializer.validated_data['request_id'],
                user=request.user,
            ),
            'messages': [],
            'confirmation_required': False,
        }
    )


class DeliveryItemMoveSerializer(Serializer, ForceCommit):
    source = IntegerField(min_value=1, source='src_request_id', default=None)
    destination = IntegerField(min_value=1, source='dest_request_id', allow_null=True)
    item = IntegerField(min_value=1, source='item_id')


@bo_api(['POST'])
def move_delivery_item(request):
    serializer = DeliveryItemMoveSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    src_request_id, dest_request_id, messages, conf_required = actions.move_delivery_item(
        user=request.user,
        **serializer.validated_data
    )
    return JsonResponse(
        {
            'source_data': (
                retrieve.get_delivery_request_detail(
                    request_id=src_request_id,
                    user=request.user,
                ) if src_request_id is not None
                else None
            ),
            'destination_data': (
                retrieve.get_delivery_request_detail(
                    request_id=dest_request_id,
                    user=request.user,
                ) if dest_request_id is not None
                else None
            ),
            'messages': messages,
            'confirmation_required': conf_required,
        }
    )


class RequestItemWorkersAddSerializer(Serializer, ForceCommit):
    request = IntegerField(min_value=1, source='request_id')
    item = IntegerField(min_value=1, source='item_id')
    workers = ListField(child=IntegerField(min_value=1), allow_empty=False, source='worker_ids')


@bo_api(['POST'])
def add_item_workers(request):
    serializer = RequestItemWorkersAddSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    _, messages, confirmation_required = actions.add_item_workers(
        user=request.user,
        **serializer.validated_data
    )
    return JsonResponse(
        {
            'data': retrieve.get_delivery_request_detail(
                request_id=serializer.validated_data['request_id'],
                user=request.user,
            ),
            'messages': messages,
            'confirmation_required': confirmation_required,
        }
    )


class RequestItemWorkerSerializer(Serializer):
    request = IntegerField(min_value=1, source='request_id')
    item = IntegerField(min_value=1, source='item_id')
    worker = IntegerField(min_value=1, source='worker_id')


class RequestItemWorkerConfirmSerializer(RequestItemWorkerSerializer, ForceCommit):
    pass


class RequestItemWorkerRemoveSerializer(RequestItemWorkerConfirmSerializer):
    reason = ChoiceField(
        choices=ItemWorkerRejection.REASON_CHOICES,
        allow_null=True,
        default=None,
    )


@bo_api(['POST'])
def remove_item_worker(request):
    serializer = RequestItemWorkerRemoveSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    _, messages, confirmation_required = actions.remove_item_worker(
        user=request.user,
        **serializer.validated_data
    )
    return JsonResponse(
        {
            'data': retrieve.get_delivery_request_detail(
                request_id=serializer.validated_data['request_id'],
                user=request.user,
            ),
            'messages': messages,
            'confirmation_required': confirmation_required,
        }
    )


@bo_api(['POST'])
def delete_item_worker(request):
    serializer = RequestItemWorkerConfirmSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    messages, confirmation_required = actions.delete_item_worker(
        user=request.user,
        **serializer.validated_data
    )
    return JsonResponse(
        {
            'data': retrieve.get_delivery_request_detail(
                request_id=serializer.validated_data['request_id'],
                user=request.user,
            ),
            'messages': messages,
            'confirmation_required': confirmation_required,
        }
    )


class ItemWorkerStartSerializer(RequestItemWorkerSerializer):
    image = ImageField(default=None)


@bo_api(['POST'])
def item_worker_start(request):
    serializer = ItemWorkerStartSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    actions.log_itemworker_start(
        user=request.user,
        location=None,
        force_commit=True,  # commit suspicious location
        **serializer.validated_data
    )
    return JsonResponse(
        {
            'data': retrieve.get_delivery_request_detail(
                request_id=serializer.validated_data['request_id'],
                user=request.user,
            ),
            'messages': [],
            'confirmation_required': False,
        }
    )


class PhotoOpSerializer(Serializer):
    request = IntegerField(min_value=1, source='request_id')
    photo = IntegerField(min_value=1, source='photo_id')


class PhotoOpCommentSerializer(PhotoOpSerializer, ForceCommit):
    comment = CharField()


@bo_api(['POST'])
def item_worker_start_reject_photo(request):
    serializer = PhotoOpCommentSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    (
        _,
        item_id,
        worker_id,
        messages,
        confirmation_required,
    ) = actions.reject_accept_itemworker_start_photo(
        user=request.user,
        is_valid=False,
        **serializer.validated_data
    )
    return JsonResponse(
        {
            'data': retrieve.get_delivery_photo_confirmation_detail(
                'start',
                item_id,
                worker_id
            ),
            'messages': messages,
            'confirmation_required': confirmation_required,
        }
    )


@bo_api(['POST'])
def item_worker_start_confirm_photo(request):
    serializer = PhotoOpSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    _, item_id, worker_id, _, _ = actions.reject_accept_itemworker_start_photo(
        user=request.user,
        is_valid=True,
        **serializer.validated_data
    )
    return JsonResponse(
        {
            'data': retrieve.get_delivery_photo_confirmation_detail(
                'start',
                item_id,
                worker_id
            ),
            'messages': [],
            'confirmation_required': False,
        }
    )


@bo_api(['POST'])
def item_worker_start_unconfirm(request):
    serializer = RequestItemWorkerConfirmSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    _, messages, confirmation_required = actions.confirm_unconfirm_itemworker_start(
        user=request.user,
        is_valid=False,
        **serializer.validated_data
    )
    return JsonResponse(
        {
            'data': retrieve.get_delivery_request_detail(
                request_id=serializer.validated_data['request_id'],
                user=request.user,
            ),
            'messages': messages,
            'confirmation_required': confirmation_required,
        }
    )


@bo_api(['POST'])
def item_worker_start_confirm(request):
    serializer = RequestItemWorkerSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    actions.confirm_unconfirm_itemworker_start(
        user=request.user,
        is_valid=True,
        **serializer.validated_data
    )
    return JsonResponse(
        {
            'data': retrieve.get_delivery_request_detail(
                request_id=serializer.validated_data['request_id'],
                user=request.user,
            ),
            'messages': [],
            'confirmation_required': False,
        }
    )


class ItemWorkerFinishSerializer(RequestItemWorkerSerializer):
    image = ImageField()


@bo_api(['POST'])
def item_worker_finish(request):
    serializer = ItemWorkerFinishSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    actions.log_itemworker_finish(
        user=request.user,
        location=None,
        **serializer.validated_data
    )
    return JsonResponse(
        {
            'data': retrieve.get_delivery_request_detail(
                request_id=serializer.validated_data['request_id'],
                user=request.user,
            ),
            'messages': [],
            'confirmation_required': False
        }
    )


@bo_api(['POST'])
def item_worker_finish_reject_photo(request):
    serializer = PhotoOpCommentSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    (
        _,
        item_id,
        worker_id,
        messages,
        confirmation_required
    ) = actions.reject_accept_itemworker_finish_photo(
        user=request.user,
        is_valid=False,
        **serializer.validated_data
    )
    return JsonResponse(
        {
            'data': retrieve.get_delivery_photo_confirmation_detail(
                'finish',
                item_id,
                worker_id
            ),
            'messages': messages,
            'confirmation_required': confirmation_required,
        }
    )


@bo_api(['POST'])
def item_worker_finish_confirm_photo(request):
    serializer = PhotoOpSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    _, item_id, worker_id, _, _ = actions.reject_accept_itemworker_finish_photo(
        user=request.user,
        is_valid=True,
        **serializer.validated_data
    )
    return JsonResponse(
        {
            'data': retrieve.get_delivery_photo_confirmation_detail(
                'finish',
                item_id,
                worker_id
            ),
            'messages': [],
            'confirmation_required': False,
        }
    )


@bo_api(['POST'])
def item_worker_finish_unconfirm(request):
    serializer = RequestItemWorkerConfirmSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    _, messages, confirmation_required = actions.confirm_unconfirm_itemworker_finish(
        user=request.user,
        is_valid=False,
        **serializer.validated_data
    )
    return JsonResponse(
        {
            'data': retrieve.get_delivery_request_detail(
                request_id=serializer.validated_data['request_id'],
                user=request.user,
            ),
            'messages': messages,
            'confirmation_required': confirmation_required,
        }
    )


class FinishConfirmSerializer(RequestItemWorkerConfirmSerializer):
    hours = DecimalField(
        min_value=0,
        max_value=24,
        max_digits=4,
        decimal_places=2,
        default=None
    )


@bo_api(['POST'])
def item_worker_finish_confirm(request):
    serializer = FinishConfirmSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    try:
        _, messages, confirmation_required = actions.confirm_unconfirm_itemworker_finish(
            user=request.user,
            is_valid=True,
            **serializer.validated_data
        )
    except HoursError as e:
        return JsonResponse(
            {
                'detail': e.detail,
                'code': e.code,
                'messages': e.messages
            },
            status=status.HTTP_409_CONFLICT
        )
    return JsonResponse(
        {
            'data': retrieve.get_delivery_request_detail(
                request_id=serializer.validated_data['request_id'],
                user=request.user,
            ),
            'messages': messages,
            'confirmation_required': confirmation_required,
        }
    )


class SetHoursSerializer(RequestWorkerSerializer, ForceCommit):
    hours = DecimalField(
        min_value=0,
        max_value=24,
        max_digits=4,
        decimal_places=2,
        default=None
    )


@bo_api(['POST'])
def set_worker_hours(request):
    serializer = SetHoursSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    try:
        _, messages, confirmation_required = actions.set_worker_hours(
            user=request.user,
            **serializer.validated_data
        )
    except HoursError as e:
        return JsonResponse(
            {
                'detail': e.detail,
                'code': e.code,
                'messages': e.messages
            },
            status=status.HTTP_409_CONFLICT
        )
    return JsonResponse(
        {
            'data': retrieve.get_delivery_request_detail(
                request_id=serializer.validated_data['request_id'],
                user=request.user,
            ),
            'messages': messages,
            'confirmation_required': confirmation_required,
        }
    )


@bo_api(['GET'])
def photos_list(request):
    return JsonResponse(
        {
            'data': retrieve.get_delivery_photo_confirmation_list(request.GET),
        }
    )


class ItemWorkerDiscrepancySerializer(RequestItemWorkerSerializer):
    comment = CharField()
    is_ok = BooleanField(allow_null=True)


# Todo: this needs a better name
class PhotosDetailQuerySerializer(Serializer):
    photo_type = ChoiceField(choices=['start', 'finish'])
    item = IntegerField(min_value=1, source='item_id')
    worker = IntegerField(min_value=1, source='worker_id')


# Todo: this needs a better name
@bo_api(['GET'])
def photos_detail(request):
    serializer = PhotosDetailQuerySerializer(data=request.GET)
    serializer.is_valid(raise_exception=True)
    return JsonResponse(
        {
            'data': retrieve.get_delivery_photo_confirmation_detail(
                **serializer.validated_data
            ),
        }
    )


@bo_api(['POST'])
def resolve_discrepancy(request):
    serializer = ItemWorkerDiscrepancySerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    actions.resolve_discrepancy(
        user=request.user,
        **serializer.validated_data
    )
    return JsonResponse(
        {
            'data': retrieve.get_delivery_request_detail(
                request_id=serializer.validated_data['request_id'],
                user=request.user,
            ),
            'messages': [],
            'confirmation_required': False,
        }
    )


@bo_api(['GET'])
def extra_photos(request):
    serializer = RequestQuerySerializer(data=request.GET)
    serializer.is_valid(raise_exception=True)
    return JsonResponse(
        {
            'data': retrieve.get_extra_photos(
                **serializer.validated_data
            ),
        }
    )


class AddPhotosSerializer(RequestQuerySerializer):
    images = ListField(child=ImageField(), allow_empty=False)


@bo_api(['POST'])
def add_extra_photos(request):
    serializer = AddPhotosSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    actions.add_extra_photos(
        **serializer.validated_data
    )
    return JsonResponse(
        {
            'data': retrieve.get_extra_photos(
                request_id=serializer.validated_data['request_id']
            ),
            'messages': [],
            'confirmation_required': False,
        }
    )


class DeliveryRequestImportSerializer(Serializer):
    customer = IntegerField(min_value=1, source='customer_id')
    requests = FileField(source='xlsx_file')


@bo_api(['POST'])
def import_requests(request):
    serializer = DeliveryRequestImportSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    try:
        actions.import_requests(
            user=request.user,
            **serializer.validated_data
        )
    except DjangoValidationError as ex:
        if ex.code == 'bad_file':
            raise ValidationError({
                'requests': ex.messages
            })
        else:
            raise
    return JsonResponse({})
