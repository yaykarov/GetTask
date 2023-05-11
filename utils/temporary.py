import datetime
import openpyxl

from django.db import transaction

from django.db.models import (
    OuterRef,
    Subquery,
)

from django.contrib.auth.models import User

from django.core.exceptions import ObjectDoesNotExist

from finance.models import Operation


from utils.date_time import (
    as_default_timezone,
    days_from_interval,
)

from the_redhuman_is import models


def export_gt_turnouts():
    the_beginning_of_time = datetime.date(day=1, month=6, year=2021)

    # 12616, 13940 and 16913 are gt account pks
    paysheet_entries = models.Paysheet_v2Entry.objects.filter(
        operation__credit__pk__in=[12616, 13940, 16913],
        paysheet__last_day__gte=the_beginning_of_time,
    ).distinct()

    turnouts = models.WorkerTurnout.objects.filter(
        turnoutoperationtopay__operation__paysheet_entry_operation__entry__in=paysheet_entries
    )

    workers = models.Worker.objects.filter(
        worker_turnouts__in=turnouts
    ).distinct(
    ).annotate(
        first_timesheet=Subquery(
            turnouts.filter(
                worker__pk=OuterRef('pk')
            ).select_related(
                'timesheet'
            ).order_by(
                'timesheet__sheet_date'
            ).values('timesheet')[:1]
        ),
        first_day=Subquery(
            models.TimeSheet.objects.filter(pk=OuterRef('first_timesheet')).values('sheet_date')
        )
    )

    gt_turnouts = models.WorkerTurnout.objects.none()

    for worker in workers:
        worker_turnouts = models.WorkerTurnout.objects.filter(
            worker=worker,
            timesheet__sheet_date__gte=max(worker.first_day, the_beginning_of_time)
        ).distinct()

        gt_turnouts = gt_turnouts.union(worker_turnouts)

    ordered_turnouts = models.WorkerTurnout.objects.filter(
        pk__in=gt_turnouts.values_list('pk', flat=True)
    ).select_related(
        'timesheet',
        'worker'
    ).order_by(
        'timesheet__sheet_date'
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    row = 1

    ws.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 10
    ws.column_dimensions[openpyxl.utils.get_column_letter(2)].width = 10
    ws.column_dimensions[openpyxl.utils.get_column_letter(3)].width = 40
    ws.column_dimensions[openpyxl.utils.get_column_letter(4)].width = 20

    for turnout in ordered_turnouts:
        col = 1

        timesheet = turnout.timesheet
        worker = turnout.worker

        ws.cell(row, col, value=timesheet.pk)
        col += 1

        cell = ws.cell(row, col, value=timesheet.sheet_date)
        cell.number_format = 'DD.MM.YY'
        col += 1

        ws.cell(row, col, value=str(worker))
        col += 1

        ws.cell(row, col, value=timesheet.customer.cust_name)
        col += 1

        ws.cell(row, col, value=timesheet.cust_location.location_name)
        col += 1

        ws.cell(row, col, value=turnout.hours_worked)
        col += 1

        worker_amount = '?'
        try:
            worker_amount = turnout.turnoutoperationtopay.operation.amount
        except ObjectDoesNotExist as e:
            pass
        ws.cell(row, col, value=worker_amount)
        col += 1

        customer_amount = '?'
        try:
            customer_amount = turnout.turnoutcustomeroperation.operation.amount
        except ObjectDoesNotExist as e:
            pass
        ws.cell(row, col, value=customer_amount)

        row += 1

    wb.save('gt_turnouts.xlsx')


@transaction.atomic
def update_customer_calculator(customer_service, first_day, customer_amount):
    calculator = models.ServiceCalculator.objects.get(
        customer_service=customer_service,
        last_day__isnull=True
    )

    calculator.last_day = first_day - datetime.timedelta(days=1)
    calculator.save()

    amount_calculator = models.clone_amount_calculator(
        calculator.calculator
    )

    interval = amount_calculator.customer_calculator.intervals.all().get()
    interval.k = customer_amount
    interval.save()

    models.ServiceCalculator.objects.create(
        customer_service=customer_service,
        calculator=amount_calculator,
        first_day=first_day
    )


def _fill_statistics_sheet(ws, first_day, last_day, count_func):
    days = days_from_interval(first_day, last_day)
    ws.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 12
    for i in range(24):
        ws.cell(i+3, 1, value=f'{i:02d}:01-{((i+1)%24):02d}:00')

    for i, day in enumerate(days):
        col = i + 2
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 6

        cell = ws.cell(1, col, value=day)
        cell.number_format = 'DD.MM'
        cell = ws.cell(2, col, value=day)
        cell.number_format = 'NN'

        values = count_func(day)

        for j, count in enumerate(values):
            cell = ws.cell(j+3, col, value=count)
            if day.weekday() > 4:
                cell.fill = openpyxl.styles.PatternFill(
                    patternType='solid',
                    fgColor=openpyxl.styles.colors.Color(rgb='00FFEEEE'),
                )


def export_gt_statistics(first_day, last_day):
    wb = openpyxl.Workbook()

    timepoint_min = Subquery(
        models.DeliveryItem.objects.filter(
            request__pk=OuterRef('pk')
        ).order_by(
            'interval_begin'
        ).values('interval_begin')[:1]
    )

    timepoint_max = Subquery(
        models.DeliveryItem.objects.filter(
            request__pk=OuterRef('pk')
        ).order_by(
            'interval_end'
        ).values('interval_end')[:1]
    )

    def _count_interval(day, timepoint_query):
        requests = models.DeliveryRequest.objects.filter(
            date=day
        ).annotate(timepoint=timepoint_query)

        counts = []
        for i in range(24):
            gt = f'{i:02d}:{00}'
            lte = f'{((i+1)%24):02d}:00'
            counts.append(
                requests.filter(
                    timepoint__gt=gt,
                    timepoint__lte=lte
                ).count()
            )

        return counts


    def _count_interval_v1(day):
        return _count_interval(day, timepoint_min)

    def _count_interval_v2(day):
        return _count_interval(day, timepoint_max)


    def _count_arrival(day):
        requests = models.DeliveryRequest.objects.filter(
            date=day
        ).annotate(
            timepoint=Subquery(
                models.ArrivalLocation.objects.filter(
                    worker__request__pk=OuterRef('pk')
                ).order_by(
                    '-timestamp'
                ).values('timestamp')[:1]
            )
        )

        counts = [0 for i in range(24)]

        for delivery_request in requests:
            if delivery_request.timepoint is None:
                continue

            delta = (
                as_default_timezone(delivery_request.timepoint) -
                as_default_timezone(
                    datetime.datetime.combine(
                        delivery_request.date,
                        datetime.time(0, 0)
                    )
                )
            )

            if delta.days > 0:
                counts[23] += 1
            else:
                hours = int(delta.seconds / 3600)
                counts[hours] += 1

        return counts

    ws = wb.active
    ws.title = 'По прибытию'
    _fill_statistics_sheet(ws, first_day, last_day, _count_arrival)

    def _add_sheet(name, func):
        ws = wb.create_sheet(name)
        _fill_statistics_sheet(ws, first_day, last_day, func)

    _add_sheet('По нижней границе времени подачи', _count_interval_v1)
    _add_sheet('По верхней границе времени подачи', _count_interval_v2)

    wb.save('gt_statistics.xlsx')


def try_reattach_photos():
    for worker in models.Worker.objects.all():
        photos = models.get_photos(worker)

        try:
            migration_photos = photos.filter(image__icontains='migration')
            if migration_photos.exists():
                migration_card, _ = models.WorkerMigrationCard.objects.get_or_create(
                    worker=worker
                )
                for photo in migration_photos:
                    photo.change_target(migration_card)

        except ObjectDoesNotExist:
            pass

        try:
            passport_photos = photos.filter(image__icontains='passport')
            if passport_photos.exists():
                passport = models.WorkerPassport.objects.get(
                    workers_id=worker,
                    is_actual=True
                )
                for photo in passport_photos:
                    photo.change_target(passport)

        except ObjectDoesNotExist:
            pass


RATES_01_04 = []

try:
    from .rates_01_04_local import *
except ImportError as e:
    print(e)


def update_tariffs_01_04():
    author = User.objects.get(username='zallexx')
    customer_pk = 21
    customer = models.Customer.objects.get(pk=customer_pk)
    last_day = datetime.date(day=2, month=4, year=2021)
    first_day = datetime.date(day=3, month=4, year=2021)

    calculators = models.ServiceCalculator.objects.filter(
        customer_service__customer=customer_pk,
        last_day__isnull=True,
    )

    for calculator in calculators:
        calculator.last_day = last_day
        calculator.save()

    zones_to_update = (
        ('mo_17', 'msk_15', 'МО до 15'),
        ('mo_31', 'msk_30', 'МО до 30'),
        ('mo_43', 'msk_60', 'МО до 60'),
        ('mo_43+', 'msk_60+', 'МО от 60'),
    )

    for old_code, new_code, new_name in zones_to_update:
        models.DeliveryZone.objects.filter(
            code=old_code
        ).update(
            code=new_code,
            name=new_name
        )

    delivery_services_to_create = (
        ('',     ''),
        ('_15',  ' до 15 км'),
        ('_30',  ' до 30 км'),
        ('_60',  ' до 60 км'),
        ('_60+', ' от 60 км'),
    )

    for zone, hours, rate_4, rate_7, rate_inf, service_name in RATES_01_04:
        if zone == 'adler':
            zone_group = 'sochi'
        else:
            zone_group = zone

        service, _ = models.Service.objects.get_or_create(name=service_name)

        try:
            customer_service = models.CustomerService.objects.get(
                customer__pk=customer_pk,
                service=service
            )

        except ObjectDoesNotExist:
            customer_service = models.create_customer_service(customer, service)

        for i, (zone_suffix, name_suffix) in enumerate(delivery_services_to_create):
            if zone == 'msk' and zone_suffix != '':
                name = 'МО'
            elif zone == 'spb':
                name = 'СПБ'
            else:
                name = service_name

            models.DeliveryService.objects.create(
                zone=zone+zone_suffix,
                service=customer_service,
                customer_service_name=service_name,
                operator_service_name=name+name_suffix,
                travel_hours=i,
                hours=hours+i,
            )

        customer_location = models.LocationZoneGroup.objects.get(
            zone_group__code=zone_group,
            location__customer_id=customer_pk
        ).location

        customer_calculator = models.SingleTurnoutCalculator.objects.create(
            parameter_1='hours',
            parameter_2='hours'
        )
        customer_calculator.intervals.add(
            models.CalculatorInterval.objects.create(
                begin=0,
                b=0,
                k=rate_4 * 1.2
            )
        )
        customer_calculator.intervals.add(
            models.CalculatorInterval.objects.create(
                begin=4.0001,
                b=0,
                k=rate_7 * 1.2
            )
        )
        customer_calculator.intervals.add(
            models.CalculatorInterval.objects.create(
                begin=7.0001,
                b=0,
                k=rate_inf * 1.2
            )
        )
        customer_calculator.save()

        amount_calculator = models.AmountCalculator.objects.create(
            customer_calculator=customer_calculator,
            foreman_calculator=models.SingleTurnoutCalculator.objects.create(
                parameter_1='hours',
                parameter_2='tariffs_01_04_21'
            ),
            worker_calculator=models.SingleTurnoutCalculator.objects.create(
                parameter_1='hours',
                parameter_2='tariffs_01_04_21'
            ),
        )

        models.ServiceCalculator.objects.create(
            customer_service=customer_service,
            calculator=amount_calculator,
            first_day=first_day
        )


def update_customer_values_01_04():
    author = User.objects.get(username='zallexx')
    customer_pk = 21
    customer = models.Customer.objects.get(pk=customer_pk)
    first_day = datetime.date(day=1, month=4, year=2021)
    last_day = datetime.date(day=2, month=4, year=2021)

    delivery_requests = models.DeliveryRequest.objects.filter(
        date__range=(first_day, last_day),
        customer=customer,
        delivery_service__isnull=False,
    )

    for delivery_request in delivery_requests:
        delivery_service = delivery_request.delivery_service

        zone = delivery_service.zone
        customer_service = delivery_request.delivery_service.service

        turnouts = models.WorkerTurnout.objects.filter(
            assignedworkerturnout__assigned_worker__request=delivery_request
        ).distinct()
        for turnout in turnouts:
            turnout.turnoutservice.customer_service = customer_service
            turnout.turnoutservice.save()

            try:
                turnout_customer_operation = models.TurnoutCustomerOperation.objects.get(
                    turnout=turnout
                )

            except ObjectDoesNotExist:
                print(f'There is no payment operation: {turnout}')

            else:
                operation = turnout_customer_operation.operation
                operation.credit = customer_service.account_76
                operation.save()

            turnout_operation = models.TurnoutOperationToPay.objects.get(turnout=turnout)

            worker = turnout.worker

            if worker.selfemployment_data.filter(deletion_ts__isnull=True).exists():
                debit = customer_service.account_20_selfemployed_work
                tax_debit = customer_service.account_20_selfemployed_taxes
            else:
                debit = customer_service.account_20_general_work
                tax_debit = customer_service.account_20_general_taxes

            Operation.objects.filter(pk=turnout_operation.operation.pk).update(debet=debit)

            tax_operation = models.TurnoutTaxOperation.objects.get(turnout=turnout)
            Operation.objects.filter(pk=tax_operation.operation.pk).update(debet=tax_debit)
