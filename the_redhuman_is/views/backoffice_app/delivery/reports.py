import itertools
import operator
from collections import defaultdict

import openpyxl
from django.db.models import (
    F,
    OuterRef,
    Q,
    Subquery,
    Value,
)
from django.http import HttpResponse
from rest_framework.fields import (
    DateField,
    IntegerField,
)
from rest_framework.serializers import Serializer

from doc_templates.http_responses import xlsx_content_response

from the_redhuman_is.geo_utils import (
    distance_to_MKAD,
    is_point_inside_MKAD,
)

from the_redhuman_is.models.models import (
    Customer,
    CustomerLocation,
)
from the_redhuman_is.models.delivery import (
    DeliveryItem,
    DeliveryRequest,
    DeliveryService,
    NormalizedAddress,
)
from the_redhuman_is.views import delivery
from the_redhuman_is.views.backoffice_app.auth import bo_api
from the_redhuman_is.views.utils import get_first_last_day
from utils import date_time
from utils.expressions import PostgresConcatWS

MSK_ZONES = ['msk', 'msk_520', 'mo_17', 'mo_31', 'mo_43', 'mo_43+']
STATUS_CODES, STATUS_LABELS = zip(*DeliveryRequest.STATUS_TYPES)


def nested(x):
    return defaultdict(lambda: nested(x-1)) if x > 0 else defaultdict(int)


def distribution_by_distance(request):
    first_day, last_day = get_first_last_day(request)

    delivery_items = DeliveryItem.objects.filter(
        Q(request__delivery_service__zone__in=MSK_ZONES) |
        Q(request__delivery_service__isnull=True),
        request__date__range=(first_day, last_day),
    ).annotate(
        latlon=Subquery(
            NormalizedAddress.objects.filter(
                location=OuterRef('pk')
            ).annotate(
                latlon=PostgresConcatWS(
                    Value(' '),
                    F('latitude'),
                    F('longitude'),
                )
            ).filter(
                version=OuterRef('address_version')
            ).values(
                'latlon'
            )[:1]
        )
    ).order_by(
        'request_id',
    ).values_list(
        'latlon',
        'request_id',
        'request__delivery_service_id',
        'request__status',
    )

    counts = nested(2)

    def _distance_to_MKAD(latlon):
        lat, lon = latlon.split(' ')
        if is_point_inside_MKAD(lat, lon):
            return 0
        return distance_to_MKAD(lat, lon)

    for _, item_group in itertools.groupby(delivery_items, key=operator.itemgetter(1)):
        first = next(item_group)
        try:
            distance = round(
                max(
                    _distance_to_MKAD(item[0])
                    for item in itertools.chain([first], item_group)
                ) / 1000
            )
        except AttributeError:
            distance = -1  # no coordinates
        counts[first[2]][distance][first[3]] += 1

    if not counts:
        return HttpResponse('nothing found')

    delivery_services = list(
        DeliveryService.objects.filter(
            pk__in=counts.keys()
        ).order_by(
            'service__customer__cust_name',
            'zone',
            'is_for_united_request',
        ).values_list(
            'pk',
            'service__customer__cust_name',
            'zone',
            'operator_service_name',
            'is_for_united_request',
        )
    )
    delivery_services.append((None, 'NA', 'NA', 'NA', False))
    count_totals = nested(1)

    for service in delivery_services:
        by_distance = counts[service[0]]
        for distance, by_status in by_distance.items():
            for status, num_requests in by_status.items():
                count_totals[distance][status] += num_requests

    wb = openpyxl.Workbook(write_only=True)
    header = ('',) + STATUS_LABELS
    ws = wb.create_sheet(title='TOTAL')
    ws.append(header)
    for distance in sorted(count_totals.keys()):
        by_status = count_totals[distance]
        row = [by_status[status] for status in STATUS_CODES]
        ws.append([distance] + row)

    for service in delivery_services:
        ws = wb.create_sheet(
            title=f'{"M|" * service[4]}{service[1]}|{service[2]}|{service[3]}'
        )
        ws.append(header)
        for distance in sorted(counts[service[0]].keys()):
            by_status = counts[service[0]][distance]
            row = [by_status[status] for status in STATUS_CODES]
            ws.append([distance] + row)

    return xlsx_content_response(
        wb,
        f'request_counts_{first_day}_{last_day}.xlsx'
    )


class DayReportSerializer(Serializer):
    day = DateField(input_formats=[date_time.DATE_FORMAT])
    customer = IntegerField(min_value=1, source='customer_id')
    location = IntegerField(min_value=1, source='location_id')


@bo_api(['GET'])
def day_report(request):
    serializer = DayReportSerializer(data=request.GET)
    serializer.is_valid(raise_exception=True)
    data = serializer.validated_data

    Customer.objects.get(pk=data['customer_id'])
    CustomerLocation.objects.get(customer_id=data['customer_id'], pk=data['location_id'])

    return delivery._day_report(**data)


class IntervalReportSerializer(Serializer):
    first_day = DateField(input_formats=[date_time.DATE_FORMAT])
    last_day = DateField(input_formats=[date_time.DATE_FORMAT])
    customer = IntegerField(min_value=1, source='customer_id')
    location = IntegerField(min_value=1, source='location_id', allow_null=True, default=None)


@bo_api(['GET'])
def interval_report(request):
    serializer = IntervalReportSerializer(data=request.GET)
    serializer.is_valid(raise_exception=True)
    data = serializer.validated_data

    customer = Customer.objects.only('pk', 'cust_name').get(
        pk=data['customer_id']
    )

    location_id = pk=data['location_id']
    if location_id is None:
        location = None
    else:
        location = CustomerLocation.objects.only('pk', 'location_name').get(
            customer_id=data['customer_id'],
            pk=location_id
        )

    return delivery.interval_report_response(
        data['first_day'],
        data['last_day'],
        customer,
        location,
    )


class DateIntervalReportSerializer(Serializer):
    first_day = DateField(input_formats=[date_time.DATE_FORMAT])
    last_day = DateField(input_formats=[date_time.DATE_FORMAT])


@bo_api(['GET'])
def suspicious_turnouts_report(request):
    serializer = DateIntervalReportSerializer(data=request.GET)
    serializer.is_valid(raise_exception=True)
    return delivery.suspicious_turnouts_report_response(
        **serializer.validated_data
    )
