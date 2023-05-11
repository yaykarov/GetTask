# -*- coding: utf-8 -*-

# Todo: remove this file

import openpyxl

from django.shortcuts import render
from django.shortcuts import redirect

from django.views.decorators.http import require_POST

from the_redhuman_is import models

from the_redhuman_is.models.customer_specific import kari

from the_redhuman_is.views.customer_specific.parser_utils import find_worker

from the_redhuman_is.services.turnout_calculations import update_turnout_payments

from utils.date_time import date_from_string

KARI_PK = 20
KARI_SERVICE_PK = 22

OPERATIONS_TO_ASSORTMENT = {
    'Собрано штук': {
        'normal': 'штука',
        'vk': 'штука ВК'
    },
    'Собрано штук с мезонина к воротам К1' : {
        'normal': 'штука с мезонина к К1'
    },
    'Собрано штук с мезонина к воротам К2' : {
        'normal': 'штука с мезонина к К2'
    },
    'Собрано штук с мезонина к воротам К3' : {
        'normal': 'штука с мезонина к К3'
    },
    'Сбор штук с мезонина на сортировку' : {
        'normal': 'штука с мезонина на сортировку'
    },
#    'Выполнено паллетных перемещений' : {
#        'normal': 'паллетное перемещение'
#    },
#    'Выполнено штучных перемещений' : {
#        'normal': 'штучное перемещение'
#    },
    'Выполнено штучных перемещений в мезонине' : {
        'normal': 'штучное перемещение в мезонине'
    },
    'Оптимизация зон хранения в мезонине' : {
        'normal': 'оптимизация зон хранения в мезонине'
    },
    'Выполнено паллетных перемещений в мезонине' : {
        'normal': 'паллетное перемещение в мезонине'
    },
    'Сортировка по магазинам' : {
        'normal': 'сортировка по магазинам'
    },
    'Закрытие ячейки' : {
        'normal': 'закрытие ячейки'
    },
}


def management_page(request):
    return render(
        request,
        'the_redhuman_is/customer_specific/kari/management_page.html',
    )


def performance_files(request):
    return render(
        request,
        'the_redhuman_is/customer_specific/kari/performance_files.html',
        {
            'files': kari.KariFile.objects.all().order_by('-pk')
        }
    )


def _date(cell):
    if cell.is_date:
        return cell.value
    else:
        return date_from_string(cell.value)


def _parse_companies(ws):
    companies = [
        'ИП САЕНКО',
    ]

    workers = {}

    for row in range(3, ws.max_row + 1):
        if ws.cell(row, 2).value is None:
            break
        company = ws.cell(row, 1).value
        if not company in companies:
            continue

        full_name = ws.cell(row, 2).value
        short_name = ws.cell(row, 3).value

        workers[short_name] = full_name

    return workers


def _parse_kari_file(f):
    wb = openpyxl.load_workbook(f, data_only=True)

    workers = _parse_companies(wb['Компании'])

    performance = {}
    def _set_performance(day, worker, assortment, count):
        if day not in performance:
            performance[day] = { worker: {} }
        if worker not in performance[day]:
            performance[day][worker] = {}

        if assortment in performance[day][worker]:
            raise Exception(
                'У работника {} за {} дублируется операция {}'.format(
                    worker,
                    day,
                    assortment
                )
            )

        performance[day][worker][assortment] = count

    def _find_worker(short_name):
        full_name = workers[short_name]
        worker = find_worker(
            KARI_PK,
            full_name
        )

        if not worker:
            raise Exception(
                'Не найден работник с именем {}, выходивший на Кари'.format(
                    full_name
                )
            )

        return worker

    wms = wb['WMS']
    for row in range(2, wms.max_row + 1):
        if wms.cell(row, 1).value is None:
            break

        short_name = wms.cell(row, 2).value
        if short_name not in workers.keys():
            continue

        operation = wms.cell(row, 3).value
        if operation not in OPERATIONS_TO_ASSORTMENT:
            continue

        worker = _find_worker(short_name)

        count = int(wms.cell(row, 5).value)
        count_vk = int(wms.cell(row, 6).value)
        day = _date(wms.cell(row, 9))

        if count > 0:
            assortment = OPERATIONS_TO_ASSORTMENT[operation]['normal']
            _set_performance(day, worker, assortment, count)

        if count_vk > 0:
            assortment = OPERATIONS_TO_ASSORTMENT[operation]['vk']
            _set_performance(day, worker, assortment, count_vk)

    movements = wb['шт. перемещения']
    for row in range(4, wms.max_row + 1):
        short_name = movements.cell(row, 1).value
        if short_name is None:
            break
        if short_name not in workers.keys():
            continue

        worker = _find_worker(short_name)
        day = _date(movements.cell(row, 5))
        count_k = int(movements.cell(row, 11).value)
        count_i = int(movements.cell(row, 9).value)

        if count_k > 0:
            _set_performance(day, worker, 'паллетное перемещение', count_k)

        if count_i > 0:
            _set_performance(day, worker, 'штучное перемещение', count_i)

    return performance


def _setup_turnouts(performance, author):
    kari_types = models.BoxType.objects.filter(
        customer__pk=KARI_PK
    )

    for day, parsed_turnouts in performance.items():
        for worker, info in parsed_turnouts.items():
            turnout = models.WorkerTurnout.objects.filter(
                timesheet__customer__pk=KARI_PK,
                timesheet__sheet_date=day,
                worker=worker,
                turnoutservice__customer_service__pk=KARI_SERVICE_PK,
            )

            if not turnout.exists():
                raise Exception(
                    'Нет выхода для {} от {}'.format(worker, day)
                )
            else:
                turnout = turnout.get()

            for box_type in kari_types:
                amount = info.get(box_type.name, 0)
                models.set_turnout_output(
                    turnout,
                    box_type,
                    amount
                )

            update_turnout_payments(turnout.pk, author)


@require_POST
def import_performance_file(request):
    f = request.FILES['performance']
    performance = _parse_kari_file(f)
    _setup_turnouts(performance, request.user)

    kari.create_kari_file(request.user, f)

    return redirect('the_redhuman_is:kari_management_page')
