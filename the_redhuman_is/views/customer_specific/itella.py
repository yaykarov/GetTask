# -*- coding: utf-8 -*-

import re
import datetime
import openpyxl

from django.shortcuts import redirect
from django.shortcuts import render

from django.db import transaction

from django.views.decorators.http import require_POST

from the_redhuman_is.auth import staff_account_required

from the_redhuman_is import forms
from the_redhuman_is import models

from the_redhuman_is.views.customer_specific.parser_utils import find_worker
from the_redhuman_is.views.customer_specific.parser_utils import find_difference


ITELLA_PK = 40
ITELLA_K2K_PKS = [61, 76, 77, 81]


@staff_account_required
def management_page(request):
    return render(
        request,
        'the_redhuman_is/itella/management_page.html',
        {
        }
    )


@require_POST
@staff_account_required
def save_k2k_sheet(request):
    f = request.FILES['k2k_sheet']
    k2k_sheet = models.itella.create_k2k_sheet(request.user, f)

    return redirect('the_redhuman_is:itella_k2k_sheet_report', pk=k2k_sheet.pk)


@staff_account_required
def k2k_sheets_list(request):
    sheets = models.itella.K2KSheet.objects.all(
    ).order_by(
        '-timestamp'
    )

    return render(
        request,
        'the_redhuman_is/itella/k2k_sheets_list.html',
        {
            'sheets': sheets
        }
    )


_K2K_SERVICE_PKS = {
    'Водитель погрузчика': 3,
    'Водитель погрузчика стажер': 3,
    'Грузчик стажер': 1,
    'Грузчик': 1,
    'Грузчик ЛИЗИНГ': 1,
    'Кладовщик': 11,
    'Кладовщик стажер': 11,
    'Комплектовщик-грузчик стажер': 2,
    'Комплектовщик стажер' : 2,
    'Комплектовщик-грузчик': 2,
    'Маркировщик ЛИЗИНГ': 9,
    'Маркировщик стажер': 9,
    'Маркировщик': 9,
}


def _parse_k2k_sheet(k2k_sheet_file):
    wb = openpyxl.load_workbook(k2k_sheet_file)
    ws = wb['Табель']

    not_found_workers = []

    turnouts = {}
    for row in range(3, ws.max_row + 1):
        day = ws.cell(row, 1).value
        full_name = ws.cell(row, 2).value
        service_name = ws.cell(row, 3).value
        if service_name not in _K2K_SERVICE_PKS:
            raise Exception(
                'Неизвестное значение в поле "Должность": {}, ячейка {}'.format(
                    service_name,
                    ws.cell(row, 3).coordinate
                )
            )
        service_pk = _K2K_SERVICE_PKS[service_name]
        hours = ws.cell(row, 4).value
        shift = ws.cell(row, 5).value
        if shift == 'НЧ':
            is_night = True
        elif shift is None:
            is_night = False
        else:
            raise Exception(
                'Неизвестное значение в поле "Смена": {}, ячейка {}'.format(
                    shift,
                    ws.cell(row, 5).coordinate
                )
            )

        worker = find_worker(
            ITELLA_PK,
            full_name
        )

        if not worker:
            if full_name not in not_found_workers:
                not_found_workers.append(full_name)
            continue

        if is_night:
            day = day - datetime.timedelta(days=1)
        day = day.date()

        if day not in turnouts:
            turnouts[day] = {}

        if worker not in turnouts[day]:
            turnouts[day][worker] = []

        turnouts[day][worker].append(
            (
                service_pk,
                hours,
                is_night
            )
        )

    return not_found_workers, turnouts


@staff_account_required
def k2k_sheet_report(request, pk):
    k2k_sheet = models.K2KSheet.objects.get(pk=pk)

    not_found_workers, file_turnouts = _parse_k2k_sheet(
        k2k_sheet.data_file
    )

    candidates_for_alias = []
    for i in range(len(not_found_workers)):
        candidates_for_alias.append(
            (
                not_found_workers[i],
                forms.WorkerSearchForm(
                    field_name='worker_{}'.format(i)
                )
            )
        )

    (
        itella_missed_days,
        itella_missed_turnouts,
        alpha_missed_days,
        alpha_missed_turnouts,
        different_hours
    ) = find_difference(file_turnouts, ITELLA_K2K_PKS, is_k2k=True)

    return render(
        request,
        'the_redhuman_is/itella/k2k_sheet_report.html',
        {
            'candidates_for_alias': candidates_for_alias,
            'itella_missed_days': itella_missed_days,
            'itella_missed_turnouts': itella_missed_turnouts,
            'alpha_missed_days': alpha_missed_days,
            'alpha_missed_turnouts': alpha_missed_turnouts,
            'different_hours': different_hours
        }
    )


@require_POST
@staff_account_required
@transaction.atomic
def save_k2k_aliases(request):
    alias_rx = re.compile('^alias_(\d+)$')

    for key, value in request.POST.items():
        m = alias_rx.match(key)
        if not m:
            continue
        worker_key = 'worker_{}'.format(m.group(1))

        if worker_key in request.POST:
            models.K2KAlias.objects.create(
                author=request.user,
                alias=value,
                worker=models.Worker.objects.get(pk=request.POST[worker_key])
            )

    return redirect('the_redhuman_is:itella_management_page')
