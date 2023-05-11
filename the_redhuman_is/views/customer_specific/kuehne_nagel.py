# -*- coding: utf-8 -*-

import openpyxl

from django.http import HttpResponse
from django.http import JsonResponse

from django.shortcuts import render
from django.shortcuts import redirect

from django.views.decorators.http import require_POST

from the_redhuman_is import forms
from the_redhuman_is import models

from the_redhuman_is.auth import staff_account_required

from the_redhuman_is.views.customer_specific.parser_utils import find_worker
from the_redhuman_is.views.customer_specific.parser_utils import find_difference


KN_PK = 45
KN_LOCATIONS = [72]


@staff_account_required
def management_page(request):
    return render(
        request,
        'the_redhuman_is/customer_specific/kuehne_nagel/management_page.html',
        {
        }
    )


@require_POST
@staff_account_required
def save_kn_sheet(request):
    f = request.FILES['kn_sheet']

    kn_sheet = models.kuehne_nagel.create_kn_sheet(request.user, f)

    return redirect('the_redhuman_is:kuehne_nagel_kn_sheet_report', pk=kn_sheet.pk)


@staff_account_required
def kn_sheets_list(request):
    sheets = models.kuehne_nagel.KNSheet.objects.all(
    ).order_by(
        '-timestamp'
    )

    return render(
        request,
        'the_redhuman_is/customer_specific/kuehne_nagel/kn_sheets_list.html',
        {
            'sheets': sheets
        }
    )


_KN_SERVICE_PKS = {
    'Водитель погрузчика': 3,
    'Грузчик': 1,
    'Комплектовщик': 2,
}


def _parse_workers(logins_sheet):
    _services_cache = {}
    def _service(service_pk):
        if service_pk not in _services_cache:
            _services_cache[service_pk] = models.Service.objects.get(
                pk=service_pk
            )
        return _services_cache[service_pk]

    not_found_workers = []
    workers = {}
    for row in range(2, logins_sheet.max_row + 1):
        code = logins_sheet.cell(row, 1).value.strip()
        if code in workers:
            raise Exception(
                'Duplicated code {}.'.format(code)
            )

        service_name_in_file = (logins_sheet.cell(row, 3).value or '').strip()

        service = _service(
            _KN_SERVICE_PKS[service_name_in_file]
        )

        full_name = logins_sheet.cell(row, 4).value.strip()
        worker = find_worker(
            KN_PK,
            full_name
        )

        if worker:
            workers[code] = (worker, service)
        else:
            not_found_workers.append(full_name)

    return workers, not_found_workers


def _parse_turnouts(workers, performance_sheet):
    turnouts = {} 

    for row in range(5, performance_sheet.max_row + 1):
        day = performance_sheet.cell(row, 1).value.date()
        code = performance_sheet.cell(row, 2).value.strip()
        shift = performance_sheet.cell(row, 4).value.strip()
        if shift == 'Night':
            is_night = True
        else:
            is_night = False
        hours = performance_sheet.cell(row, 6).value

        if code not in workers:
            # Todo: some error?
            continue

        worker, service = workers[code]

        if day not in turnouts:
            turnouts[day] = {}

        if worker not in turnouts[day]:
            turnouts[day][worker] = []

        turnouts[day][worker].append(
            (
                service.pk,
                hours,
                is_night
            )
        )

    return turnouts


def _parse_kn_sheet(kn_sheet_file):
    from the_redhuman_is.views.fine_utils import _find_worker

    wb = openpyxl.load_workbook(kn_sheet_file)

    workers, not_found_workers = _parse_workers(
        wb['СKVRTR - Логины']
    )

    turnouts = _parse_turnouts(workers, wb['Продуктивность'])

    return not_found_workers, turnouts


@staff_account_required
def kn_sheet_report(request, pk):
    kn_sheet = models.KNSheet.objects.get(pk=pk)

    not_found_workers, turnouts = _parse_kn_sheet(kn_sheet.data_file)

    candidates_for_alias = []
    for i in range(len(not_found_workers)):
        candidates_for_alias.append(
            (
                not_found_workers[i],
                forms.WorkerSearchForm(
                    field_name='worker_{}'.format(i)
                )
            )
        )

    (
        itella_missed_days,
        itella_missed_turnouts,
        alpha_missed_days,
        alpha_missed_turnouts,
        different_hours
    ) = find_difference(turnouts, KN_LOCATIONS)

    return render(
        request,
        'the_redhuman_is/itella/k2k_sheet_report.html',
        {
            'candidates_for_alias': candidates_for_alias,
            'itella_missed_days': itella_missed_days,
            'itella_missed_turnouts': itella_missed_turnouts,
            'alpha_missed_days': alpha_missed_days,
            'alpha_missed_turnouts': alpha_missed_turnouts,
            'different_hours': different_hours
        }
    )
