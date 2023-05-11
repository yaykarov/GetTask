import datetime
import enum
import openpyxl

from dataclasses import dataclass
from decimal import Decimal

from typing import (
    List,
    Tuple,
)

from django.contrib.auth.models import User

from django.db import transaction
from django.db.models import QuerySet

from django.utils import timezone

from the_redhuman_is.async_utils.talk_bank.exceptions import (
    TalkBankExceptionError
)
from the_redhuman_is.async_utils.talk_bank.talk_bank import (
    SelfemploymentsStatus,
    Service,
    TalkBankClient as TalkBankClientApi,
    create_talk_bank_client,
)

from the_redhuman_is.models import (
    PaysheetEntryTalkBankPayment,
    PaysheetEntryTalkBankPaymentAttempt,
    PaysheetEntryTalkBankIncomeRegistration,
    PaysheetTalkBankPaymentStatus,
    Paysheet_v2,
    Paysheet_v2Entry,
    TalkBankBindStatus,
    TalkBankClient,
    Worker,
    WorkerReceipt,
    WorkerReceiptPaysheetEntry,
)

from .common import get_service_name


class WorkerOperationResult(enum.Enum):
    ok = 'ok'
    client_creation_failed = 'client_creation_failed'
    client_unregistered = 'client_unregistered'
    client_unbound = 'client_unbound'
    client_status_error = 'client_status_error'
    client_bind_requested = 'client_bind_requested'
    client_bind_failed = 'client_bind_failed'
    client_already_bound = 'client_already_bound'
    income_registration_failed = 'income_registration_failed'
    payment_failed = 'payment_failed'
    invoice_cancellation_failed = 'invoice_cancellation_failed'


class BindWorkerContext():
    def __init__(
            self,
            bank_client: TalkBankClientApi,
            # Worker should be annotated with
            # with_talk_bank_client_id()
            # with_passport()
            # with_selfemployment_data()
            worker: Worker,
    ):
        self.bank_client = bank_client
        self.worker = worker

    def bank_client_exists(self) -> bool:
        return self.worker.talk_bank_client_id != None

    def get_selfemployment_status(self) -> Tuple[SelfemploymentsStatus, str]:
        return self.bank_client.get_selfemployment_status(self.worker.talk_bank_client_id)

    @transaction.atomic
    def create_bank_client(self) -> None:
        if self.worker.wse is None:
            raise Exception(f'Нет данных о самозанятости для {self.worker}')

        self.worker.talk_bank_client_id = str(self.worker.pk)
        TalkBankClient.objects.create(
            worker=self.worker,
            client_id=self.worker.talk_bank_client_id
        )
        self.bank_client.create_client_with_simple_identification(
            client_id=self.worker.talk_bank_client_id,
            first_name=self.worker.name,
            middle_name=self.worker.patronymic,
            last_name=self.worker.last_name,
            birth_day=self.worker.birth_date.strftime('%Y-%m-%d'),
            gender=0 if self.worker.sex == 'Муж' else 1,
            secret_word=f'{self.worker.talk_bank_client_id}-nTXDXxgdQLwvLR9f',
            phone=self.worker.tel_number,
            document_serial=self.worker.passport['series'],
            document_number=self.worker.passport['number'],
            document_date=self.worker.passport['date_of_issue'],
            inn=self.worker.wse['tax_number'],
        )

    def bind_client(self):
        return self.bank_client.bind_client(
            client_id=self.worker.talk_bank_client_id,
            raise_exception=True
        )

    def save_bind_result(self, result: WorkerOperationResult, description: str) -> None:
        do_save_bind_result(self.worker.pk, result, description)


class WorkerContext(BindWorkerContext):
    def __init__(
            self,
            author: User,
            bank_client: TalkBankClientApi,
            # Worker should be annotated with
            # with_talk_bank_client_id()
            # with_passport()
            # with_selfemployment_data()
            # with_full_name()
            worker: Worker,
            amount: Decimal,
            service_description: str,
            day: datetime.date,
            paysheet_id: int
    ):
        super().__init__(bank_client, worker)
        self.author = author
        self.amount = amount
        self.service_description = service_description
        self.day = day
        self.paysheet_id = paysheet_id

    def register_income(self) -> None:
        income_registration_request_id: str = self.bank_client.register_income_from_legal_entity(
            client_id=self.worker.talk_bank_client_id,
            operation_time=self.day.strftime('%Y-%m-%dT08:00:00+03:00'),
            services=[Service(name=self.service_description, amount=self.amount, quantity=1)],
            total_amount=self.amount,
            customer_inn='5029258192',
            customer_organization='ОБЩЕСТВО С ОГРАНИЧЕННОЙ ОТВЕТСТВЕННОСТЬЮ "ГЕТТАСК"',
        )
        PaysheetEntryTalkBankIncomeRegistration.objects.create(
            author=self.author,
            paysheet_id=self.paysheet_id,
            worker=self.worker,
            income_registration_request=income_registration_request_id,
            date=self.day,
            amount=self.amount,
        )

    def make_payment(self) -> None:
        transfer_result = self.bank_client.transfer(
            amount=self.amount,
            account=self.worker.wse['bank_account'],
            bik=self.worker.wse['bank_identification_code'],
            name=self.worker.full_name,
            inn=self.worker.wse['tax_number'],
            description=self.service_description,
        )

        # Todo: what if transfer is ok but model creation is not?

        PaysheetEntryTalkBankPayment.objects.create(
            paysheet_entry=Paysheet_v2Entry.objects.get(paysheet_id=self.paysheet_id, worker=self.worker), # Todo!
            tax_number=self.worker.wse['tax_number'],
            amount=self.amount,
            bank_account=self.worker.wse['bank_account'],
            bank_name=self.worker.wse['bank_name'],
            bank_identification_code=self.worker.wse['bank_identification_code'],
            completed=transfer_result.completed,
            status=transfer_result.status,
            order_slug=transfer_result.order_slug,
            talk_bank_commission=transfer_result.talk_bank_commission,
            beneficiary_partner_commission=transfer_result.beneficiary_partner_commission,
        )

    def cancel_invoice(self):
        # Todo: also delete a record about invoice
        pass


def do_save_bind_result(worker_id: int, result: WorkerOperationResult, description: str):
    TalkBankBindStatus.objects.create(
        worker_id=worker_id,
        is_bound=(result == WorkerOperationResult.client_already_bound),
        operation_result=result.value,
        operation_result_description=description
    )


def do_bind_worker(context: BindWorkerContext) -> Tuple[WorkerOperationResult, str]:
    if not context.bank_client_exists():
        try:
            context.create_bank_client()
        except TalkBankExceptionError as error:
            return WorkerOperationResult.client_creation_failed, error.message
        except Exception as e:
            return WorkerOperationResult.client_creation_failed, str(e)

    try:
        status, description = context.get_selfemployment_status()
    except TalkBankExceptionError as error:
        return WorkerOperationResult.client_status_error, error.message
    except Exception as e:
        return WorkerOperationResult.client_status_error, str(e)

    if status == SelfemploymentsStatus.registered:
        return WorkerOperationResult.client_already_bound, 'Работник уже привязан к ТокБанку'
    else:
        if status == SelfemploymentsStatus.unregistered:
            return WorkerOperationResult.client_unregistered, description
        elif status == SelfemploymentsStatus.unbound:
            try:
                context.bind_client()
                return WorkerOperationResult.client_bind_requested, description
            except TalkBankExceptionError as error:
                return WorkerOperationResult.client_bind_failed, error.message
            except Exception as e:
                return WorkerOperationResult.client_bind_failed, str(e)
        else:
            return WorkerOperationResult.client_status_error, description


def bind_worker(context: BindWorkerContext) -> Tuple[WorkerOperationResult, str]:
    result, description = do_bind_worker(context)
    context.save_bind_result(result, description)
    return result, description


def register_worker_income_async(context: WorkerContext) -> Tuple[WorkerOperationResult, str]:
    result, description = bind_worker(context)
    if result == WorkerOperationResult.client_already_bound:
        pass
    elif result == WorkerOperationResult.client_bind_requested:
        return WorkerOperationResult.client_unbound, description
    else:
        return result, description

    try:
        context.register_income()
        return WorkerOperationResult.ok, 'Доход зарегистрирован успешно'
    except TalkBankExceptionError as error:
        return WorkerOperationResult.income_registration_failed, error.message
    except Exception as e:
        return WorkerOperationResult.income_registration_failed, str(e)


def _get_talk_bank_client_and_related_data(
        tb_client_id: str,
        income_registration_request_id: str,
) -> Tuple[TalkBankClient, PaysheetEntryTalkBankIncomeRegistration, Paysheet_v2Entry]:
    talk_bank_client: TalkBankClient = TalkBankClient.objects.get(client_id=tb_client_id)
    income_registration: PaysheetEntryTalkBankIncomeRegistration = PaysheetEntryTalkBankIncomeRegistration.objects.get(
        worker=talk_bank_client.worker,
        income_registration_request=income_registration_request_id,
    )
    paysheet_entry: Paysheet_v2Entry = Paysheet_v2Entry.objects.get(
        paysheet=income_registration.paysheet,
        worker=talk_bank_client.worker,
    )
    return talk_bank_client, income_registration, paysheet_entry


def do_register_talkbank_income(
        tb_client_id: str,
        income_registration_request_id: str,
        receipt_url: str
) -> Tuple[User, Worker, Paysheet_v2]:
    talk_bank_client, income_registration, paysheet_entry = _get_talk_bank_client_and_related_data(
        tb_client_id=tb_client_id, income_registration_request_id=income_registration_request_id
    )

    with transaction.atomic():
        receipt = WorkerReceipt.objects.create(
            author=income_registration.author,
            worker=talk_bank_client.worker,
            url=receipt_url,
            date=income_registration.date,
        )
        WorkerReceiptPaysheetEntry.objects.create(
            worker_receipt=receipt,
            paysheet_entry=paysheet_entry,
        )

    return income_registration.author, talk_bank_client.worker, income_registration.paysheet


def do_make_worker_payment(context: WorkerContext) -> Tuple[WorkerOperationResult, str]:
    try:
        context.make_payment()
    except Exception as e:
        try:
            context.cancel_invoice()
        except TalkBankExceptionError as error:
            return WorkerOperationResult.invoice_cancellation_failed, error.message
        except Exception as error:
            return WorkerOperationResult.invoice_cancellation_failed, str(error)

        if isinstance(e, TalkBankExceptionError):
            return WorkerOperationResult.payment_failed, e.message
        else:
            return WorkerOperationResult.payment_failed, str(e)

    return WorkerOperationResult.ok, 'Оплачено успешно'


def on_talkbank_income_registered(tb_client_id: str, income_registration_request_id: str, receipt_url: str) -> None:
    author, worker, paysheet = do_register_talkbank_income(tb_client_id, income_registration_request_id, receipt_url)

    worker = _paysheet_workers_with_all_stuff(paysheet).get(pk=worker.pk)
    service_description = get_service_name(
        worker.paysheet_workdays['start'],
        worker.paysheet_workdays['end']
    )
    payment_day = timezone.localdate(paysheet.timestamp)

    bank_client = create_talk_bank_client()

    context = WorkerContext(
        author=author,
        bank_client=bank_client,
        worker=worker,
        amount=Decimal(worker.paysheet_entry['amount']),
        service_description=service_description,
        day=payment_day,
        paysheet_id=paysheet.pk
    )
    payment_status, description = do_make_worker_payment(context)
    PaysheetEntryTalkBankPaymentAttempt.objects.create(
        paysheet_id=paysheet.pk,
        worker=worker,
        status=payment_status,
        description=description,
        service_description=service_description,
    )

    if payment_status != WorkerOperationResult.ok:
        try:
            paysheet_entry = Paysheet_v2Entry.objects.get(pk=worker.paysheet_entry['pk'])
            paysheet_entry.delete()
        except Exception:
            pass

    close_paysheet_if_ready(paysheet_id=paysheet.pk)


def on_talkbank_income_registration_failed(tb_client_id: str, income_registration_request_id: str, errors: str):
    talk_bank_client, income_registration, paysheet_entry = _get_talk_bank_client_and_related_data(
        tb_client_id=tb_client_id, income_registration_request_id=income_registration_request_id
    )

    PaysheetEntryTalkBankPaymentAttempt.objects.create(
        paysheet_id=paysheet_entry.paysheet.pk,
        worker=talk_bank_client.worker,
        status=WorkerOperationResult.income_registration_failed,
        description=errors,
        service_description='',
    )

    try:
        paysheet_entry.delete()
    except Exception:
        pass

    close_paysheet_if_ready(paysheet_id=paysheet_entry.paysheet.pk)


@transaction.atomic()
def close_paysheet_if_ready(paysheet_id: int):
    paysheet: Paysheet_v2 = Paysheet_v2.objects.select_for_update().get(pk=paysheet_id)
    not_ready_entries: QuerySet[Paysheet_v2Entry] = Paysheet_v2Entry.objects.filter(
        paysheet=paysheet
    ).with_has_successful_talkbank_payment(
    ).exclude(
        has_successful_talkbank_payment=True
    )
    if not not_ready_entries.exists():
        status: PaysheetTalkBankPaymentStatus = PaysheetTalkBankPaymentStatus.objects.get(paysheet=paysheet)
        status.status = PaysheetTalkBankPaymentStatus.COMPLETE
        status.save(update_fields=['status'])
        paysheet.close_with_default_payment_account(author=status.author)


@dataclass
class WorkerPaymentReport:
    paysheet_id: int
    worker_id: int
    full_name: str
    tax_number: str
    amount: Decimal
    account: str
    bank_name: str
    bank_id: str # БИК
    receipt_url: str
    receipt_description: str
    payment_status: str
    payment_result: str
    payment_result_description: str


def make_payment_report_xlsx(reports: List[WorkerPaymentReport]):
    wb = openpyxl.Workbook(write_only=True)

    ws = wb.create_sheet()

    ws.page_setup.paperSize = openpyxl.worksheet.worksheet.Worksheet.PAPERSIZE_A4
    ws.page_setup.orientation = openpyxl.worksheet.worksheet.Worksheet.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    captions = [
        ('ID ведомости', 15),
        ('ID работника', 15),
        ('ФИО работника', 35),
        ('ИНН', 17),
        ('Сумма', 10),
        ('Банковский счет', 23),
        ('Банк', 35),
        ('БИК', 14),
        ('Чек', 45),
        ('Услуга', 65),
        ('Статус оплаты', 20),
        ('Результат выплаты', 40),
        ('Описание', 65),
    ]

    for column, width in enumerate((caption[1] for caption in captions), start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(column)].width = width

    def _make_cell(value):
        cell = openpyxl.cell.WriteOnlyCell(ws, value=value)
        cell.alignment = openpyxl.styles.Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True,
        )
        return cell

    ws.append([_make_cell(caption[0]) for caption in captions])

    for report in reports:
        values = [
            report.paysheet_id,
            report.worker_id,
            report.full_name,
            report.tax_number,
            report.amount,
            report.account,
            report.bank_name,
            report.bank_id,
            report.receipt_url,
            report.receipt_description,
            report.payment_status,
            report.payment_result,
            report.payment_result_description,
        ]
        ws.append([_make_cell(value) for value in values])

    return wb


def get_paysheet_payment_report(paysheet_id: int):
    workers = Worker.objects.filter_by_payment_attempt(
        paysheet_id=paysheet_id
    ).with_paysheet_workdays(
        paysheet_id=paysheet_id
    ).with_full_name(
    ).with_talk_bank_payment_attempt(
        paysheet_id=paysheet_id
    ).with_talk_bank_payment(
        paysheet_id=paysheet_id
    ).with_paysheet_receipt_url(
        paysheet_id=paysheet_id
    )

    def _talk_bank_payment(worker, key):
        if worker.talk_bank_payment is None:
            return '-'
        return worker.talk_bank_payment[key]

    reports = []
    for worker in workers:
        reports.append(
            WorkerPaymentReport(
                paysheet_id=paysheet_id,
                worker_id=worker.id,
                full_name=worker.full_name,
                tax_number=_talk_bank_payment(worker, 'tax_number'),
                amount=_talk_bank_payment(worker, 'amount'),
                account=_talk_bank_payment(worker, 'bank_account'),
                bank_name=_talk_bank_payment(worker, 'bank_name'),
                bank_id=_talk_bank_payment(worker, 'bank_identification_code'),
                receipt_url=worker.paysheet_receipt_url,
                receipt_description=worker.talk_bank_payment_attempt['service_desciption'],
                payment_status=_talk_bank_payment(worker, 'status'),
                payment_result=worker.talk_bank_payment_attempt['payment_result'],
                payment_result_description=worker.talk_bank_payment_attempt['payment_result_description'],
            )
        )

    return make_payment_report_xlsx(reports)


def _paysheet_workers_with_all_stuff(paysheet: Paysheet_v2) -> QuerySet[Worker]:
    return paysheet.workers(
    ).with_paysheet_workdays(
        paysheet_id=paysheet.pk
    ).with_paysheet_entry(
        paysheet_id=paysheet.pk
    ).with_talk_bank_client_id(
    ).with_passport(
    ).with_selfemployment_data(
        more=True,
    ).with_full_name(
    ).with_is_selfemployed(
    ).with_cardholder_name(
    ).with_worker_type(
    )


@transaction.atomic
def do_start_paysheet_payments(author_id: int, paysheet_id: int) -> None:
    author = User.objects.get(pk=author_id)
    paysheet = Paysheet_v2.objects.get(pk=paysheet_id)

    workers: QuerySet[Worker] = _paysheet_workers_with_all_stuff(paysheet)
    payment_day = timezone.localdate(paysheet.timestamp)

    bank_client = create_talk_bank_client()
    bank_client.enable_logging()

    for worker in workers:
        service_description = None
        if worker.worker_type == 'selfemployed_own_account':
            service_description = get_service_name(
                worker.paysheet_workdays['start'],
                worker.paysheet_workdays['end']
            )
            context = WorkerContext(
                author=author,
                bank_client=bank_client,
                worker=worker,
                amount=Decimal(worker.paysheet_entry['amount']),
                service_description=service_description,
                day=payment_day,
                paysheet_id=paysheet_id
            )

            income_registration_status, income_registration_status_description = register_worker_income_async(context)
        else:
            income_registration_status = 'wrong_worker_type'
            income_registration_status_description = f'Неподходящий для выплаты через ТокБанк "тип" работника: {worker.worker_type}'

        if income_registration_status != WorkerOperationResult.ok:
            PaysheetEntryTalkBankPaymentAttempt.objects.create(
                paysheet_id=paysheet_id,
                worker=worker,
                status=income_registration_status,
                description=income_registration_status_description,
                service_description=service_description,
            )

            try:
                paysheet_entry = Paysheet_v2Entry.objects.get(pk=worker.paysheet_entry['pk'])
                paysheet_entry.delete()
            except Exception:
                pass

    transaction.on_commit(lambda: close_paysheet_if_ready(paysheet_id=paysheet_id))


def start_paysheet_payments(author_id: int, paysheet_id: int) -> None:
    with transaction.atomic():
        paysheet = Paysheet_v2.objects.select_for_update(
            nowait=True
        ).get(
            pk=paysheet_id
        )
        if hasattr(paysheet, 'paysheettalkbankpaymentstatus'):
            return

        paysheet_payment_status = PaysheetTalkBankPaymentStatus.objects.create(
            author_id=author_id,
            paysheet=paysheet,
        )

    try:
        do_start_paysheet_payments(author_id, paysheet_id)
    except Exception:
        paysheet_payment_status.status = PaysheetTalkBankPaymentStatus.ERROR

    paysheet_payment_status.save()


def bind_worker_sync(worker_id: int) -> None:
    worker = Worker.objects.all(
    ).with_talk_bank_client_id(
    ).with_passport(
    ).with_full_name(
    ).with_selfemployment_data(
        more=True,
    ).get(
        pk=worker_id
    )

    bank_client=create_talk_bank_client()
    bank_client.enable_logging()

    bind_worker(
        BindWorkerContext(
            bank_client=bank_client,
            worker=worker
        )
    )
