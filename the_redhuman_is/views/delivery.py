# -*- coding: utf-8 -*-
import csv
import datetime
import io
import itertools
import json
import operator
import os
import random
from collections import defaultdict
from decimal import Decimal

import openpyxl
import openpyxl.cell
import openpyxl.styles
import openpyxl.utils
import openpyxl.worksheet
import openpyxl.worksheet.worksheet
import openpyxl.writer
import openpyxl.writer.excel

from functools import reduce
from operator import add

from zipfile import ZipFile

from crispy_forms.helper import FormHelper
from crispy_forms.layout import (
    Field,
    Layout,
    Row,
)
from django.conf import settings
from django.contrib.auth import authenticate
from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User
from django.contrib.contenttypes.models import ContentType
from django.contrib.postgres.aggregates import (
    ArrayAgg,
    StringAgg,
)
from django.contrib.postgres.fields import ArrayField

from django.core.exceptions import (
    ObjectDoesNotExist,
    ValidationError,
)
from django.urls import reverse

from django.db import transaction

from django.db.models import (
    Case,
    CharField,
    Count,
    DateTimeField,
    DecimalField,
    Exists,
    F,
    FloatField,
    IntegerField,
    JSONField,
    Max,
    Min,
    OuterRef,
    Q,
    Subquery,
    Sum,
    Value,
    When,
)

from django.db.models.functions import (
    Cast,
    Ceil,
    Coalesce,
    JSONObject,
    NullIf,
)

from django.shortcuts import (
    redirect,
    render,
)

from django.views.decorators.http import (
    require_GET,
    require_POST,
)
from django.views.decorators.csrf import csrf_exempt

from django.http import (
    HttpResponse,
    HttpResponseForbidden,
    HttpResponseNotFound,
    JsonResponse,
)

from django.utils import timezone

from rest_framework.decorators import api_view
from rest_framework_simplejwt.tokens import RefreshToken

from dal.autocomplete import Select2QuerySetView

from urllib.parse import quote

from doc_templates.http_responses import xlsx_content_response

from the_redhuman_is import forms
from the_redhuman_is import models
from the_redhuman_is.models.models import (
    TimeSheet,
    WorkerTurnout,
    get_user_location,
)
from the_redhuman_is.models.delivery import (
    DailyReconciliation,
    DeliveryItem,
    DeliveryRequest,
    DeliveryRequestOperator,
    DeliveryZone,
    ItemWorker,
    ItemWorkerFinish,
    ItemWorkerStart,
    LastLocation,
    Location,
    MobileAppStatus,
    NormalizedAddress,
    OnlineStatusMark,
    OperatorZoneGroup,
    RequestWorker,
    RequestWorkerTurnout,
    ZoneGroup,
)
from the_redhuman_is.models.paysheet_v2 import (
    PayoutRequest,
    Paysheet_v2EntryOperation,
    WorkerReceipt,
)
from the_redhuman_is.models.photo import Photo
from the_redhuman_is.models.turnout_operations import TurnoutOperationToPay
from the_redhuman_is.models.worker import (
    MobileAppWorker,
    Worker,
    WorkerComments,
    WorkerRating,
    WorkerUser,
)
from the_redhuman_is.services.delivery import (
    actions,
    notifications,
    tariffs,
    utils,
)
from the_redhuman_is.services.delivery_requests import (
    update_fcm_token,
    update_import_processed_timestamp,
)

from the_redhuman_is.services.delivery_utils import (
    slots_chain,
    slot_starts,
)
from the_redhuman_is.services.turnout_calculations import update_not_paid_turnout_payments
from the_redhuman_is.services.worker import (
    NoWorkPermit,
    update_online_status,
)
from the_redhuman_is.tasks import (
    normalize_address_in_bulk,
    send_sms,
    send_tg_message_to_clerks,
)

from the_redhuman_is.views.utils import (
    _get_value,
    content_response,
    exception_to_500,
    get_first_last_day,
)

from the_redhuman_is import geo_utils

from utils import excel_import as xls
from utils.date_time import (
    as_default_timezone,
    date_from_string,
    days_from_interval,
    postgres_str_to_datetime,
    string_from_date,
    time_interval_format,
)
from utils.expressions import (
    Haversine,
    PostgresConcatWS,
)
from utils.files import jpeg_to_jpg
from utils.forms import SubmitNoValue
from utils.numbers import (
    ZERO_OO,
    lcm,
)
from utils.functools import (
    partition,
    strtobool,
)
from utils.phone import (
    format_phone,
    normalized_phone,
)


class DataIntegrityException(Exception):
    pass


def _get_first_last_day(request):
    first_day, last_day = get_first_last_day(request, set_initial=False)
    if not first_day:
        first_day = timezone.now().date() - datetime.timedelta(days=1)
    if not last_day:
        last_day = timezone.now().date() + datetime.timedelta(days=1)

    return first_day, last_day


# ### Autocomplete


def locations_qs(view):
    customer_pk = view.forwarded.get('customer')
    if not customer_pk:
        return models.CustomerLocation.objects.none()
    locations = models.CustomerLocation.objects.filter(
        customer_id__pk=customer_pk
    ).order_by('location_name')
    if view.q:
        locations = locations.filter(location_name__icontains=view.q)
    return locations


class LocationAutocomplete(LoginRequiredMixin, Select2QuerySetView):
    def get_queryset(self):
        return locations_qs(self)

    def get_result_label(self, result):
        return result.location_name


def services_qs(view):
    customer_pk = view.forwarded.get('customer')
    delivery_request_pk = view.forwarded.get('delivery_request_pk')
    force_allowed_services = [2, 9]

    if not customer_pk:
        return models.DeliveryService.objects.none()

    services = models.DeliveryService.objects.filter(
        # Todo: temporary remove "big mass" services
        Q(min_mass=0) | Q(pk__in=force_allowed_services),
        service__active=True,
        service__customer__pk=customer_pk,
    )
    if view.q:
        services = services.filter(
            operator_service_name__icontains=view.q
        )

    if delivery_request_pk:
        delivery_request = DeliveryRequest.objects.with_items_count(
        ).get(
            pk=delivery_request_pk
        )
        is_united = delivery_request.items_count > 1

        # God forgive me...
        if customer_pk == 21:
            pivot = datetime.date(day=3, month=4, year=2021)
            if delivery_request.date < pivot:
                services = services.exclude(pk__gt=1000)
                services = services.filter(
                    is_for_united_request=is_united
                )
            else:
                services = services.exclude(pk__lte=1000)
        else:
            services = services.filter(
                is_for_united_request=is_united
            )

        if delivery_request.location:
            groups = models.LocationZoneGroup.objects.filter(
                location__customer_id__pk=customer_pk
            )
            if groups.exists():
                zones = DeliveryZone.objects.filter(
                    group__locationzonegroup__location=delivery_request.location
                ).values_list('code', flat=True)
                services = services.filter(
                    zone__in=zones
                )

    return services.order_by(
        'zone',
        'min_mass',
    )


class ServiceAutocomplete(LoginRequiredMixin, Select2QuerySetView):
    def get_queryset(self):
        return services_qs(self)

    def get_result_label(self, result):
        return result.operator_service_name


def delivery_requests_qs(view):
    customer_pk = view.forwarded.get('customer')
    date = date_from_string(view.forwarded.get('date'))
    delivery_requests = DeliveryRequest.objects.filter(
        customer__pk=customer_pk,
        date=date,
    ).order_by('pk')
    if view.q:
        delivery_requests = delivery_requests.annotate(
            pk_str=Cast('pk', CharField())
        ).filter(
            Q(route__icontains=view.q) |
            Q(pk_str__icontains=view.q)
        )
    return delivery_requests


class DeliveryRequestAutocomplete(LoginRequiredMixin, Select2QuerySetView):
    def get_queryset(self):
        return delivery_requests_qs(self)

    def get_result_label(self, result):
        if result.route:
            return f'{result.route} ({result.pk})'
        else:
            return f'{result.pk}'


def delivery_items_qs(view):
    delivery_request = DeliveryRequest.objects.get(pk=view.forwarded.get('request'))
    delivery_items = DeliveryItem.objects.filter(
        request__date=delivery_request.date,
        request__customer=delivery_request.customer_id,
        request__driver_name=delivery_request.driver_name,
        request__driver_phones=delivery_request.driver_phones,
    ).exclude(
        request=delivery_request
    ).select_related(
        'request'
    ).order_by(
        'pk'
    )
    if view.q:
        delivery_items = delivery_items.filter(
            Q(code__icontains=view.q) |
            Q(address__icontains=view.q)
        )
    return delivery_items


class DeliveryItemAutocomplete(LoginRequiredMixin, Select2QuerySetView):
    def get_queryset(self):
        return delivery_items_qs(self)

    def get_result_label(self, result):
        return '{} {}'.format(result.code, result.address)


def delivery_workers(view):
    workers = Worker.objects.all(
    ).filter_mobile(
    ).filter(
        banned__isnull=True,
    ).filter_rf_or_mk(
    ).select_related(
        'workertag'
    )

    delivery_request_pk = view.forwarded.get('delivery_request_pk')
    if delivery_request_pk:
        delivery_request = DeliveryRequest.objects.get(
            pk=delivery_request_pk
        )
        today = timezone.localdate()
        workers = workers.annotate(
            distance=Case(
                When(
                    lastlocation__location__timestamp__date__gte=today,
                    then=Subquery(
                        NormalizedAddress.objects.filter(
                            version=F('location__address_version'),
                            location__request=delivery_request_pk
                        ).annotate(
                            distance=Haversine(
                                'latitude',
                                OuterRef('lastlocation__location__latitude'),
                                'longitude',
                                OuterRef('lastlocation__location__longitude'),
                            ),
                        ).values(
                            'location__request'
                        ).annotate(
                            min_distance=Min('distance')
                        ).values(
                            'min_distance'
                        )
                    )
                ),
                default=None,
                output_field=FloatField()
            )
        ).order_by(
            'distance'
        )

        location = delivery_request.location_id
        if location is not None:
            try:
                zone_group = models.LocationZoneGroup.objects.get(
                    location=location
                ).zone_group_id
                workers = workers.filter(
                    Q(workerzone__isnull=True) |
                    Q(workerzone__zone=zone_group)
                )
            except models.LocationZoneGroup.DoesNotExist:
                pass

    if view.q:
        workers = workers.filter_by_text(view.q)
    return workers


class DeliveryWorkerAutocomplete(LoginRequiredMixin, Select2QuerySetView):
    def get_queryset(self):
        return delivery_workers(self)

    def get_result_label(self, result):
        try:
            comment = f' [{result.workercomments.text}]'
        except WorkerComments.DoesNotExist:
            comment = ''
        distance = getattr(result, 'distance', None)
        if distance is not None and round(distance) <= DISTANCE_LIMIT:
            distance = f' {round(distance)} км'
        else:
            distance = ' - км'
        return f'{result}{comment}{distance}'


def operators(view):
    first_day = date_from_string(view.forwarded.get('first_day'))
    last_day = date_from_string(view.forwarded.get('last_day'))

    users = User.objects.order_by(
        'pk'
    )

    if view.q:
        users = users.filter(
            Q(username__icontains=view.q) |
            Q(first_name__icontains=view.q)
        )

    operator_qs = DeliveryRequestOperator.objects.filter(
        operator=OuterRef('pk')
    )
    if first_day:
        operator_qs = operator_qs.exclude(request__date__lt=first_day)
    if last_day:
        operator_qs = operator_qs.exclude(request__date__gt=last_day)

    users = users.annotate(
        requests_exist=Exists(operator_qs)
    ).filter(
        requests_exist=True
    )

    return users


class DeliveryRequestOperatorAutocomplete(LoginRequiredMixin, Select2QuerySetView):
    def get_queryset(self):
        return operators(self)

    def get_result_label(self, result):
        return result.first_name or result.username


# Todo: get rid of this in favor of the same class from backoffice_app
class DeliveryZoneAutocomplete(LoginRequiredMixin, Select2QuerySetView):
    def get_queryset(self):
        qs = ZoneGroup.objects.order_by('name')
        if self.q:
            return qs.filter(name__icontains=self.q)
        return qs


class DeliveryOperatorAutocomplete(LoginRequiredMixin, Select2QuerySetView):
    def get_queryset(self):
        qs = User.objects.filter(
            is_active=True,
            groups__name='Доставка-диспетчер'
        )
        if self.q:
            return qs.filter(
                Q(first_name__icontains=self.q) |
                Q(last_name__icontains=self.q) |
                Q(username__icontains=self.q)
            )
        return qs.order_by('last_name', 'first_name')

    def get_result_label(self, result):
        return f'{result.first_name} {result.last_name}'


# Serialization utils

def _get_size(filefield):
    units = ['Б', 'К', 'М', 'Г', 'Т', 'П']  # Нуачо...
    if filefield.storage.exists(filefield.name):
        size = os.path.getsize(filefield.path)
        i = 0
        while size > 1024 and i < (len(units) - 1):
            size /= 1024
            i += 1

        return '{} {}'.format(round(size, 1), units[i])

    return 0


def _serialize_requests_file(requests_file, base_url):
    WEIGHTS = {
        'processing': 1,
        'error': 2,
        'finished_with_errors': 3,
        'finished': 4,
    }

    report = None
    if requests_file.processed_data_file:
        report_name = requests_file.processed_data_file.name
        report = {
            'name': os.path.basename(report_name),
            'size': _get_size(requests_file.processed_data_file),
            'url': base_url + '?type=processed&pk={}'.format(requests_file.pk)
        }
    name = requests_file.data_file.name

    return {
        'timepoint': as_default_timezone(requests_file.timestamp).strftime('%d.%m.%Y %H:%M'),
        'author': {
            'id': requests_file.author_id,
            'name': requests_file.author.first_name,
        },
        'imported_file': {
            'name': os.path.basename(name),
            'size': _get_size(requests_file.data_file),
            'url': base_url + '?type=imported&pk={}'.format(requests_file.pk)
        },
        'report': report,
        'status': {
            'id': requests_file.status,
            'weight': WEIGHTS.get(requests_file.status, 0),
            'text': requests_file.status_description
        }
    }


DISTANCE_LIMIT = 99


# ### XLS imports


class RequestExcelParser(xls.Parser):
    date = xls.DateField('Дата')
    route = xls.CharField(['Номер серии заявок', 'Номер маршрута'], default=None)
    code = xls.CharField(['Индекс груза', 'Груз'])
    mass = xls.FloatField(['Масса', 'Вес (кг)'])
    volume = xls.FloatField(['Объем', 'Объем (м3)'])
    place_count = xls.IntegerField(['Кол-во мест', 'Кол-во (мест)'])
    shipment_type = xls.CharField('Характер груза')
    time_interval = xls.TimeIntervalField(
        ['Интервал выполнения (напр. 10:00-16:00)', 'Интервал забора/доставки'],
    )
    address = xls.CharField('Адрес')
    driver_name = xls.CharField(
        ['ФИО водителя', 'Водитель'],
        formatter=utils.normalize_driver_name,
        default=''
    )
    driver_phones = xls.CharField(
        ['Телефоны водителя', 'Номер телефона'],
        formatter=lambda v: ','.join(utils.normalize_driver_phones(v)),
        default=''
    )
    max_size = xls.FloatField('Макс. габарит (м)', required=False, default=None)
    has_elevator = xls.BooleanTextField(['Лифт (да/нет)', 'Лифт'], required=False, default=True)
    floor = xls.IntegerField('Этаж', required=False, default=None)
    carrying_distance = xls.IntegerField('Пронос (м)', required=False, default=None)
    workers_required = xls.IntegerField(
        'Количество грузчиков поставщика',
        required=False,
        default=1,
    )


# Todo: когда появятся юниттесты, это вот одно из первых мест, которые стоит протестировать
#
# Todo: кажется, тут максимальная жесть с т.з. конкурентности и многопоточности
# пока у нас очередь из 1 процесса - как-бы все равно, но надо как-то разобраться
# Идея - как-то лочить импорт на уровне "клиента", это должно позволить упорядочить
# все модификации
@transaction.atomic
def _do_import_requests(user, customer, wb, ws, merge=True):
    location = get_user_location(user)
    codes = defaultdict(set)
    duplicated_codes = defaultdict(set)
    routes_with_format_errors = set()
    routes = defaultdict(lambda: defaultdict(list))
    ws_data = []

    parser = RequestExcelParser(ws)

    _date_location_reconciliation = {}

    def _has_confirmed_daily_reconciliation(date, location_id):
        date_location = date, location_id
        if date_location not in _date_location_reconciliation:
            _date_location_reconciliation[date_location] = DailyReconciliation.objects.filter(
                date=date,
                location=location_id,
                dailyreconciliationconfirmation__isnull=False
            ).exists()
        return _date_location_reconciliation[date_location]

    # First row is for titles
    for idx, result in enumerate(parser.parse_rows()):
        ws_data.append(result)
        data, errors = result

        if 'date' in errors or 'route' in errors:
            continue

        date = data['date']
        route = data['route']

        if route:
            routes[date][route].append(result)

        if location is not None:
            if _has_confirmed_daily_reconciliation(date, location.pk):
                errors['date'].append('Имеется подтвержденная ежедневная сверка')

        if errors:
            routes_with_format_errors.add(route)
        elif 'code' not in errors:
            code = data['code']
            if code in codes[date]:
                duplicated_codes[date].add(code)
            else:
                codes[date].add(code)

    # Check if routes are OK in imported file and in DB
    for date, day_routes in routes.items():
        for route, items in day_routes.items():

            route_errors = []

            if len(items) < 2:
                route_errors.append(
                    'В маршруте №{} меньше двух заявок, должно быть не меньше'.format(route)
                )

            data, errors = items[0]
            if 'driver_name' in errors or 'driver_phone' in errors:
                continue

            driver_name = data['driver_name']
            driver_phones = data['driver_phones']

            delivery_requests = DeliveryRequest.objects.select_for_update().filter(
                customer=customer,
                date=date,
                route=route
            )
            delivery_request_count = delivery_requests.count()

            if delivery_request_count == 1:
                delivery_request = delivery_requests.get()

                if delivery_request.driver_name != driver_name:
                    route_errors.append(
                        'В базе у маршрута №{} другой водитель ({}). '
                        'ФИО должны быть строго одинаковыми'.format(
                            route,
                            delivery_request.driver_name
                        )
                    )

                if delivery_request.driver_phones != driver_phones:
                    route_errors.append(
                        'В базе у маршрута №{} другой телефон водителя ({}). '
                        'Телефоны должны быть строго одинаковыми'.format(
                            route,
                            delivery_request.driver_phones
                        )
                    )

            elif delivery_request_count > 1:
                route_errors.append(
                    'В базе уже есть несколько маршрутов №{} за {}'.format(
                        route,
                        string_from_date(date)
                    )
                )

            for item in items[1:]:
                if driver_name != item.data['driver_name']:
                    route_errors.append(
                        'В файле у маршрута №{} есть заявки с разными водителями '
                        '(столбец {}). ФИО должны быть строго одинаковыми'.format(
                            route,
                            openpyxl.utils.get_column_letter(
                                parser.columns['driver_name'].index
                            )
                        )
                    )
                if driver_phones != item.data['driver_phones']:
                    route_errors.append(
                        'В маршруте №{} есть заявки с разными телефонами водителей '
                        '(столбец {}). Телефоны должны быть строго одинаковыми'.format(
                            route,
                            openpyxl.utils.get_column_letter(
                                parser.columns['driver_phones'].index
                            )
                        )
                    )

            for item in items:
                code = item.data['code']
                try:
                    duplicate_item = DeliveryItem.objects.get(
                        request__customer=customer,
                        request__date=date,
                        code=code,
                    )
                except DeliveryItem.DoesNotExist:
                    pass
                except DeliveryItem.MultipleObjectsReturned:
                    route_errors.append(
                        'В базе уже есть заявки, включающих индекс груза {} на дату {}'.format(
                            code,
                            string_from_date(date)
                        )
                    )
                else:
                    if duplicate_item.request.route != route:
                        if duplicate_item.request.requestworker_set.exists():
                            route_errors.append(
                                'В базе индекс {} включен в маршрут №{}, для которого уже '
                                'назначены рабочие. В файле этот индекс включен в маршрут '
                                '№{}. Автокорректировка в таком случае '
                                'не поддерживается'.format(
                                    code,
                                    duplicate_item.request.route,
                                    route
                                )
                            )

            if route_errors:
                for item in items:
                    item.errors['route'].extend(route_errors)

    for item in ws_data:
        data, errors = item
        if 'date' in errors:
            continue

        if not errors and data['route'] and (data['route'] in routes_with_format_errors):
            errors['route'].append(
                'Есть ошибки в других частях маршрута №{}'.format(
                    data['route']
                )
            )

        if 'code' not in errors and data['code'] in duplicated_codes.get(data['date'], {}):
            errors['code'].append(
                'Индекс груза {} за дату {} должен быть уникален, но встречается '
                'в файле несколько раз'.format(
                    data['code'],
                    string_from_date(data['date'])
                )
            )

    # Now all (almost) pre-checks should be OK, and we just should create some objects

    # Как это работает:
    # 1. Если есть номер маршрута - плевать на merge, просто пытаемся взять готовую заявку
    # с таким маршрутом (или создаем, если нет)
    # 2. Если нет номера маршрута и есть флаг merge, пытаемся взять подходящую существующую
    # заявку из базы или из кэша (там могут быть заявки, для которых были явно заданы номера
    # маршрутов). Отсюда, кстати, вытекает необходимость сначала создавать "маршрутные" заявки.
    # 3. Если ни номера маршрута, ни флага merge - просто создаем заявку, минуя кэш
    created_requests = {}

    def _get_or_create_delivery_request(date, route, driver_name, driver_phones):
        # Assumes (location, date) has no confirmed reconciliations
        if route or merge:
            if (date, driver_name, driver_phones) not in created_requests:
                created_requests[(date, driver_name, driver_phones)] = {}

            cached_requests = created_requests[(date, driver_name, driver_phones)]

            delivery_request = cached_requests.get(route)

            if delivery_request is None:
                if route:
                    delivery_request_qs = DeliveryRequest.objects.select_for_update(
                        of=('self',)
                    ).filter(
                        customer=customer,
                        date=date,
                        route=route
                    )
                    if delivery_request_qs.exists():
                        delivery_request = delivery_request_qs.get()

                else:  # (route is None and merge)
                    # Пока только претендент. Более подходящие могут быть в кэше.
                    delivery_request = utils.get_request_to_merge(
                        customer,
                        date,
                        driver_name,
                        driver_phones,
                        location
                    )
                    # Пытаемся выбрать заявку с максимальным номером маршрута
                    # (для консистентности с механизмом внутри get_request_to_merge)
                    for key, value in cached_requests.items():
                        if delivery_request:
                            if key > delivery_request.route:
                                delivery_request = value
                        else:
                            delivery_request = value

            if route is None:  # И значит точно merge == True
                if delivery_request is not None and delivery_request.route is None:
                    # Т.к. мы взяли существующую заявку - в данный момент у нее уже есть
                    # 1 адрес, и точно добавится еще 1. Самое время установить номер маршрута.
                    delivery_request.route = utils.free_route_number(
                        delivery_request.customer,
                        delivery_request.date
                    )
                    delivery_request.save(user=user)

            if delivery_request is None:
                delivery_request = utils.create_private_request(
                    author=user,
                    customer=customer,
                    date=date,
                    route=route,
                    driver_name=driver_name,
                    driver_phones=driver_phones,
                    location=location,
                )

            cached_requests[route] = delivery_request

            return delivery_request

        else:
            return utils.create_private_request(
                author=user,
                customer=customer,
                date=date,
                driver_name=driver_name,
                driver_phones=driver_phones,
                location=location,
            )

    routes_to_check = {}
    items_to_normalize = defaultdict(list)
    duplications = set()

    # First - requests with route number
    order = list(itertools.chain.from_iterable(
        partition(
            lambda x: ws_data[x][0]['route'] is None,
            range(len(ws_data))
        )
    ))

    _REQUEST_FIELDS = {'date', 'route', 'driver_name', 'driver_phones'}

    for idx in order:
        data, errors = ws_data[idx]

        if not errors:
            request_values = {
                key: data[key]
                for key in _REQUEST_FIELDS
            }
            item_values = {
                key: value
                for key, value in data.items()
                if key not in _REQUEST_FIELDS
            }
            interval_begin, interval_end = item_values.pop('time_interval')
            item_values['interval_begin'] = interval_begin
            item_values['interval_end'] = interval_end
            # GT-553
            item_values['has_elevator'] = True

            items = DeliveryItem.objects.select_for_update().filter(
                request__customer=customer,
                request__date=data['date'],
                code=data['code'],
            )
            items_count = items.count()
            if items_count == 0:
                delivery_request = _get_or_create_delivery_request(
                    **request_values
                )
                item = DeliveryItem.objects.create(
                    request=delivery_request,
                    **item_values,
                )
                items_to_normalize[delivery_request.pk].append(item.pk)
            else:
                if data['route']:
                    # Предполагаем, что тут все хорошо
                    # Т.е. для маршрутов - заявка всего одна, рабочие не назначены
                    # (это проверялось выше)
                    # И можно просто "переприцепить" существующий айтем к новому маршруту
                    item = items.get()
                    # На случай, если в маршруте теперь 1 или меньше заявок
                    routes_to_check[item.request.pk] = True
                    item.request = _get_or_create_delivery_request(
                        **request_values
                    )
                    item.save(update_fields=['request'])
                    if item.request.pk not in routes_to_check:
                        routes_to_check[item.request.pk] = False
                else:
                    duplications.add(idx)

    # Todo: кажется, этот блок тупо не работает (мы пытаемся делать проверки внутри атомика)
    for pk, do_ensure in routes_to_check.items():
        delivery_request = DeliveryRequest.objects.get(pk=pk)
        if do_ensure:
            delivery_request = _ensure_can_be_route(delivery_request, user=user)
        if delivery_request:
            tariffs.try_to_update_tariff(delivery_request, user)

    error_count = 0

    source_column_count = parser.source_column_count
    note_free_column = source_column_count + 1

    date_style = None
    if 'gt-date' not in wb.named_styles:
        date_style = openpyxl.styles.NamedStyle(
            name='gt-date',
            number_format='DD.MM.YY'
        )

    def _set_background(row, rgb):
        for i in range(source_column_count + 1):
            cell = ws.cell(row=row, column=i+1)
            if i > 0:
                cell.style = 'Normal'
            color = openpyxl.styles.colors.Color(rgb=rgb)
            cell.fill = openpyxl.styles.PatternFill(
                patternType='solid',
                fgColor=color,
            )

    # First row is for titles
    for row_num, result in enumerate(ws_data, start=2):

        data, errors = result

        if 'date' in data and date_style:
            cell = ws.cell(
                row=row_num,
                column=parser.columns['date'].index + 1,
                value=data['date']
            )
            cell.style = date_style

        if errors:
            error_count += 1
            _set_background(row_num, '00FFEEEE')
            ws.cell(
                row=row_num,
                column=note_free_column,
                value='; '.join(
                    msg
                    for field_errors in errors.values()
                    for msg in field_errors
                )
            )

        elif row_num - 2 in duplications:
            _set_background(row_num, '00FFFFCC')
            ws.cell(row=row_num, column=note_free_column, value='Дубликат')

        else:
            _set_background(row_num, '00EEFFEE')
            ws.cell(row=row_num, column=note_free_column, value='Импортировано')

    return items_to_normalize, len(duplications), error_count


# Todo: maybe it should be places in models or somewhere else
# !!! Should be a part of a huey task (see tasks.py)
def do_import_requests_and_make_report(user_pk, customer_pk, file_pk, notify_dispatchers):
    user = User.objects.get(pk=user_pk)
    customer = models.Customer.objects.get(pk=customer_pk)
    requests_file = models.RequestsFile.objects.get(pk=file_pk)

    wb = openpyxl.load_workbook(requests_file.data_file)
    ws = wb.active

    try:
        items_to_normalize, duplicated, errors = _do_import_requests(user, customer, wb, ws)

    except Exception as e:
        print(e)
        requests_file.status = 'error'
        if isinstance(e, ValidationError) and e.args:
            requests_file.status_description = e.args[0]
        else:
            requests_file.status_description = 'Что-то пошло не так'

    else:
        if errors > 0:
            requests_file.status = 'finished_with_errors'
        else:
            requests_file.status = 'finished'

        status = ''

        if errors > 0:
            status += 'Ошибок: {}'.format(errors)

        imported = sum(len(v) for v in items_to_normalize.values())
        if imported > 0 or (errors == 0 and duplicated == 0):
            if status:
                status += ', '
            status += 'Импортировано: {}'.format(imported)

        if duplicated > 0:
            if status:
                status += ', '
            status += 'Дублей: {}'.format(duplicated)

        requests_file.status_description = status

        for request_id, item_ids in items_to_normalize.items():
            normalize_address_in_bulk(item_ids, request_id, 0, user)

    proxy_file = io.BytesIO()
    wb.save(proxy_file)
    requests_file.processed_data_file.save(
        os.path.basename(requests_file.data_file.name),
        proxy_file
    )
    requests_file.save()

    update_import_processed_timestamp(customer)

    if notify_dispatchers:
        notifications.notify_new_import(customer)


# Ensure if delivery_request still can be a route
def _ensure_can_be_route(delivery_request, user):
    items = DeliveryItem.objects.filter(request=delivery_request)
    items_count = items.count()
    if items_count == 0:
        models.DriverSms.objects.filter(request=delivery_request).delete()
        delivery_request.delete()
        return None
    elif items_count == 1:
        delivery_request.route = None
        delivery_request.save(user=user)

    return delivery_request


def requests_file_response(requests_file_pk, file_type, customer=None):
    if file_type not in ['imported', 'processed']:
        return HttpResponse(status=404)

    try:
        params = {'pk': requests_file_pk}
        if customer:
            params['customer'] = customer
        requests_file = models.RequestsFile.objects.get(**params)

        if file_type == 'imported':
            data_file = requests_file.data_file
        else:
            data_file = requests_file.processed_data_file

        return content_response(data_file)

    except ObjectDoesNotExist as e:
        return HttpResponse(status=404)


# Dispatcher's pages


def delivery_request_extra_photos(request, pk):
    delivery_request = DeliveryRequest.objects.get(pk=pk)

    return render(
        request,
        'the_redhuman_is/reports/photos_list.html',
        {
            'photos': models.get_photos(delivery_request)
        }
    )


# ### XLS reports (day/interval)

def _get_worker_subquery():
    return RequestWorker.objects.filter(
        request=OuterRef('pk'),
        workerrejection__isnull=True,
        requestworkerturnout__isnull=False,
    ).annotate(
        worker_fields=JSONObject(
            full_name=PostgresConcatWS(
                Value(' '),
                F('worker__last_name'),
                F('worker__name'),
                NullIf(F('worker__patronymic'), Value(''))
            ),
            hours='requestworkerturnout__workerturnout__hours_worked',
            amount='requestworkerturnout__workerturnout__turnoutcustomeroperation__operation__amount',
            timesheet='requestworkerturnout__workerturnout__timesheet',
        )
    )


def tuple_to_range(row1, col1, row2, col2):
    return (
        f'{openpyxl.utils.get_column_letter(col1)}{row1}:'
        f'{openpyxl.utils.get_column_letter(col2)}{row2}'
    )


alignment_center = openpyxl.styles.Alignment(
    horizontal='center',
    vertical='center',
    wrap_text=True,
)
alignment_vertical_center = openpyxl.styles.Alignment(
    vertical='center',
)
alignment_bottom_center = openpyxl.styles.Alignment(
    horizontal='center',
)
alignment_bottom_left = openpyxl.styles.Alignment(
    horizontal='left',
    wrap_text=False,
)

border_thin = openpyxl.styles.borders.Border(
    left=openpyxl.styles.borders.Side(style='thin'),
    right=openpyxl.styles.borders.Side(style='thin'),
    top=openpyxl.styles.borders.Side(style='thin'),
    bottom=openpyxl.styles.borders.Side(style='thin'),
)
border_bottom = openpyxl.styles.borders.Border(
    bottom=openpyxl.styles.borders.Side(style='thin'),
)

font_calibri_10 = openpyxl.styles.Font(name='Calibri', size='10')
font_calibri_bold_10 = openpyxl.styles.Font(name='Calibri', size='10', bold=True)
font_superscript = openpyxl.styles.Font(vertAlign='superscript')

cell_fill_lightgreen = openpyxl.styles.PatternFill('solid', fgColor='EEFFEE')
cell_fill_lightblue = openpyxl.styles.PatternFill('solid', fgColor='EEEEFF')


def _write_interval_req_row(
        ws,
        current_row,
        values,
        alignment=None,
        border=None,
        font=None,
        fill=None,
        number_formatting=None  # 1-based numbering
):

    if number_formatting is None:
        number_formatting = {}

    def _make_cell(caption, order_num):
        cell = openpyxl.cell.WriteOnlyCell(ws, value=caption)
        cell.alignment = alignment
        cell.border = border
        if font is not None:
            cell.font = font
        if fill is not None:
            cell.fill = fill
        try:
            cell.number_format = number_formatting[order_num]
        except KeyError:
            pass
        return cell

    blank_cell = _make_cell(None, None)

    values = [v if isinstance(v, list) else [v] for v in values]
    lines_total = lcm(*(len(v) for v in values))
    cell_height = [lines_total // len(v) for v in values]

    for i_down in range(lines_total):
        row = []
        for j_across, value in enumerate(values):
            k_value, remainder = divmod(i_down, cell_height[j_across])
            has_content = remainder == 0
            if has_content:
                row.append(_make_cell(value[k_value], j_across + 1))
                if cell_height[j_across] > 1:
                    ws.merged_cells.ranges.append(tuple_to_range(
                        current_row,
                        j_across + 1,
                        current_row + cell_height[j_across] - 1,
                        j_across + 1
                    ))
            else:
                # in MS Excel, merged cell border is determined by all constituent cells
                row.append(blank_cell)
        ws.append(row)
        current_row += 1

    return current_row


def _fill_details_caption(ws):
    captions = [
        ('Дата', 12),
        ('Индекс', 15),
        ('Масса', 10),
        ('Время погрузки/выгрузки', 15),
        ('ФИО водителя', 30),
        ('Адрес', 40),
        ('Кол-во грузчиков', 10),
        ('Фио грузчиков', 40),
        ('Кол-во человеко-часов', 4),
        (None, 7),
        ('Тариф, р/ч', 10),
        ('Сумма', 10),
        ('№ фото реестра выполненных работ', 20),
        ('Отмена клиента', 16),
        ('Примечание', 20),
    ]

    ws.row_dimensions[1].height = 32

    for column, caption in enumerate(captions, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(column)].width = caption[1]

    def _make_cell(value):
        cell = openpyxl.cell.WriteOnlyCell(ws, value=value)
        cell.alignment = alignment_center
        return cell

    ws.append([_make_cell(caption[0]) for caption in captions])
    ws.merged_cells.ranges.append(tuple_to_range(1, 9, 1, 10))


def _get_day_report_queryset(day, customer_id, location_id):
    return DeliveryRequest.objects.filter(
        date=day,
        status__in=DeliveryRequest.SUCCESS_STATUSES,
        customer=customer_id,
        location=location_id
    ).with_items_count(
    ).annotate(
        item_data=Subquery(
            DeliveryItem.objects.filter(
                request=OuterRef('pk')
            ).order_by(
            ).values(
                'request'
            ).annotate(
                total_mass=Sum('mass'),
                total_code=StringAgg('code', delimiter='; '),
                begin=Min('interval_begin'),
                end=Max('interval_end'),
                total_address=StringAgg('address', delimiter='; '),
            ).annotate(
                fields=JSONObject(
                    mass='total_mass',
                    code='total_code',
                    begin='begin',
                    end='end',
                    address='total_address',
                )
            ).values(
                'fields'
            ),
            output_field=JSONField()
        ),
        workers=Coalesce(
            Subquery(
                _get_worker_subquery().values(
                    'request'
                ).annotate(
                    workers_list=ArrayAgg(
                        'worker_fields',
                        ordering=(
                            'worker__last_name',
                            'worker__name',
                            'worker__patronymic',
                            'worker__id',
                        )
                    )
                ).values(
                    'workers_list'
                )
            ),
            [],
            output_field=ArrayField(JSONField())
        )
    ).order_by(
        'items_count',
        'pk'
    ).values(
        'pk',
        'status',
        'comment',
        'driver_name',
        'route',
        'items_count',
        'item_data',
        'workers',
    )


def _fill_details_sheet(ws, day, customer_id, location_id):
    ws.title = 'Детализация'
    _fill_details_caption(ws)

    requests = _get_day_report_queryset(day, customer_id, location_id)

    total_workers = 0
    total_hours = 0
    total_amount = 0

    def _get_request_code(req):
        if req['items_count'] == 1:
            return req['item_data']['code']
        else:
            return f"{req['route']}: {req['item_data']['code']}"

    current_row = 2

    for request in requests:
        worked_hours = sum(worker['hours'] for worker in request['workers'])
        amount = sum(worker['amount'] for worker in request['workers'])
        current_row = _write_interval_req_row(
            ws,
            current_row,
            [
                day,
                _get_request_code(request),
                request['item_data']['mass'],
                # Todo: is it really the correct approach for request interval?
                '{} - {}'.format(
                    request['item_data']['begin'][:5],
                    request['item_data']['end'][:5],
                ),
                request['driver_name'],
                request['item_data']['address'],
                len(request['workers']),
                [worker['full_name'] for worker in request['workers']],
                [worker['hours'] for worker in request['workers']],
                worked_hours,
                [
                    worker['amount'] / worker['hours']
                    for worker in request['workers']
                ],
                amount,
                (
                    [worker['timesheet'] for worker in request['workers']]
                    if request['status'] == DeliveryRequest.COMPLETE
                    else [None] * len(request['workers'])
                ),
                (
                    'Отмена клиентом'
                    if request['status'] == DeliveryRequest.CANCELLED_WITH_PAYMENT
                    else None
                ),
                request['comment'],
            ],
            alignment=alignment_vertical_center,
            fill=cell_fill_lightblue if request['items_count'] > 1 else None,
            number_formatting={
                1: 'DD.MM.YYYY',
                11: '0.00',
            }
        )

        total_workers += len(request['workers'])
        total_hours += worked_hours
        total_amount += amount

    ws.append(
        [None] * 5 +
        [
            openpyxl.cell.WriteOnlyCell(ws, value=value)
            for value in ['ИТОГО:', total_workers, None, None, total_hours, None, total_amount]
        ]
    )

    return total_workers, total_hours, total_amount


def _fill_total_sheet(ws, day, workers, hours, amount):
    ws.title = 'Итог'

    for column in range(1, 6):
        ws.column_dimensions[openpyxl.utils.get_column_letter(column)].width = 15

    def _make_cell(value, alignment=None, border=None, font=None, number_format=None):
        cell = openpyxl.cell.WriteOnlyCell(ws, value=value)
        cell.alignment = alignment
        cell.border = border
        if font is not None:
            cell.font = font
        if number_format is not None:
            cell.number_format = number_format
        return cell

    captions = [
        'Дата',
        'Кол-во груз-ков',
        'Кол-во часов',
        'Сумма',
    ]

    ws.append([
        _make_cell(value, alignment=alignment_bottom_left, border=border_thin)
        for value in captions
    ])
    values_cells = [
        _make_cell(value, None, border=border_thin)
        for value in [workers, hours, amount]
    ]
    ws.append([_make_cell(
        day,
        alignment_bottom_left,
        border=border_thin,
        number_format='DD.MM.YYYY'
    )] + values_cells)
    ws.append(
        [_make_cell('Общий итог:', alignment_bottom_left, border=border_thin)] + values_cells
    )
    for _ in range(7):
        ws.append([])

    signature_lines = [
        _make_cell(None, border=border_bottom),
        _make_cell(None, border=border_bottom),
        _make_cell('/', border=border_bottom),
    ]
    signature_labels = [
        None,
        _make_cell('должность', font=font_superscript, alignment=alignment_bottom_center),
        _make_cell('подпись', font=font_superscript, alignment=alignment_bottom_center),
        _make_cell('расшифровка', font=font_superscript, alignment=alignment_bottom_center)
    ]
    ws.append([_make_cell('Исполнитель:', alignment_bottom_left)] + signature_lines)
    ws.append(signature_labels)

    for _ in range(4):
        ws.append([])

    ws.append([_make_cell('Заказчик:', alignment_bottom_left)] + signature_lines)
    ws.append(signature_labels)


def _add_images_to_zip(zip_file, day, customer_id):
    # alternatively, _fill_details_sheet could be refactored to return timesheet pks

    turnout_sq = RequestWorkerTurnout.objects.filter(
        workerturnout__timesheet=OuterRef('pk'),
        requestworker__request__date=day,
        requestworker__request__status=DeliveryRequest.COMPLETE,
        requestworker__request__customer=customer_id
    )

    timesheets = TimeSheet.objects.annotate(
        for_day_and_customer=Exists(
            turnout_sq
        )
    ).filter(
        for_day_and_customer=True
    ).annotate(
        photo_url=Subquery(
            Photo.objects.filter(
                content_type=ContentType.objects.get_for_model(TimeSheet),
                object_id=OuterRef('pk')
            ).order_by(
                '-pk'
            ).values(
                'image'
            )[:1],
            output_field=CharField()
        )
    ).filter(
        photo_url__isnull=False
    ).values(
        'pk',
        'photo_url',
    ).order_by(
        'pk'
    )
    for timesheet in timesheets:
        _, ext = os.path.splitext(timesheet['photo_url'])
        with open(
                Photo.image.field.storage.path(timesheet['photo_url']),
                'rb'
        ) as image_contents:
            zip_file.writestr(
                '{}{}'.format(timesheet['pk'], jpeg_to_jpg(ext)),
                image_contents.read()
            )


def _day_report(day, customer_id, location_id):
    wb = openpyxl.Workbook(write_only=True)
    workers, hours, amount = _fill_details_sheet(
        wb.create_sheet(),
        day,
        customer_id,
        location_id,
    )
    _fill_total_sheet(wb.create_sheet(), day, workers, hours, amount)

    # Todo: move to some utils?
    proxy_file = io.BytesIO()
    with ZipFile(proxy_file, "w") as zip_file:
        zip_file.writestr(
            'report_{}.xlsx'.format(string_from_date(day)),
            openpyxl.writer.excel.save_virtual_workbook(wb)
        )
        _add_images_to_zip(zip_file, day, customer_id)

    response = HttpResponse(content_type='application/zip')
    response[
        'Content-Disposition'
    ] = "attachment; filename*=UTF-8''{}".format(
        'day_report_{}.zip'.format(string_from_date(day))
    )
    response.write(proxy_file.getvalue())

    return response


def _fill_interval_titles(ws, current_row, first_day, last_day):
    def _make_cell(value):
        cell = openpyxl.cell.WriteOnlyCell(ws, value=value)
        cell.alignment = alignment_center
        cell.border = border_thin
        cell.font = font_calibri_bold_10
        return cell

    double_height_captions = [
        ('Дата оказания услуг', 8),
        ('Мар-т', 6),
        ('Индекс(ы)', 10),
        ('Объект оказания услуг (адрес)', 30),
        ('Время начала оказания услуг', 12),
        ('Кол-во чел-к', 5),
        ('Всего часов', 5),
        ('ФИО грузчика', 20),
        ('Кол-во часов', 5),
        ('Тариф, р/ч', 8),
        ('Сумма', 8),
        (None, 8),
        ('Кол-во чел-к', 5),
        ('Всего часов', 5),
        ('Согласованное время подачи', 13),
        ('Время начала оказания услуг', 13),
    ]

    blank_cell = openpyxl.cell.WriteOnlyCell(ws)
    blank_cell.border = border_thin

    column_widths = [caption[1] for caption in double_height_captions] + [25, 15]
    for column, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(column)].width = width
    ws.row_dimensions[3].height = 28

    ws.append(
        [_make_cell(
            f'Период с {string_from_date(first_day)} по {string_from_date(last_day)}'
        )] +
        [blank_cell] * 17
    )
    current_row += 1

    ws.append(
        [_make_cell('По заявке')] +
        [blank_cell] * 6 +
        [_make_cell('По факту')] +
        [blank_cell] * 10
    )
    ws.merged_cells.ranges.extend(
        (
            tuple_to_range(1, 1, 1, 18),
            tuple_to_range(2, 1, 2, 7),
            tuple_to_range(2, 8, 2, 18),
        )
    )
    current_row += 1

    new_row = []
    for i, caption in enumerate(double_height_captions, start=1):
        new_row.append(_make_cell(caption[0]))
        if i == 11:
            continue
        elif i == 12:
            ws.merged_cells.ranges.append(tuple_to_range(3, i - 1, 4, i))
        else:
            ws.merged_cells.ranges.append(tuple_to_range(3, i, 4, i))
    new_row.extend((
        _make_cell('Контактные данные принимающего лица - представителя Заказчика'),
        blank_cell
    ))
    ws.merged_cells.ranges.append(tuple_to_range(3, 17, 3, 18))
    ws.append(new_row)
    current_row += 1

    ws.append(
        [blank_cell] * 16 +
        [
            _make_cell('ФИО'),
            _make_cell('Телефон')
        ]
    )
    current_row += 1

    return current_row


def _fill_interval_sheet(ws, request_qs, first_day, last_day):
    # Two passes over the queryset are needed: one to set dimensions and one to write cell data.
    # Optimized write only mode only allows sequential writing to the file (no random access).
    # Sheet parameters such as column width and row height need to be set before any cell data.
    # Setting sheet parameters after WriteOnlyWorksheet.append has been called has no effect.
    current_row = 5
    for day, day_requests in itertools.groupby(request_qs, key=operator.itemgetter('date')):
        for _ in range(sum(
                lcm(len(request['items']), len(request['workers']))
                for request in day_requests
        )):
            ws.row_dimensions[current_row].height = 28
            current_row += 1
        current_row += 1
    ws.row_dimensions[current_row + 2].height = 80

    current_row = 1
    current_row = _fill_interval_titles(ws, current_row, first_day, last_day)

    total_workers = 0
    total_hours = 0
    total_amount = 0

    for day, day_requests in itertools.groupby(request_qs, key=operator.itemgetter('date')):
        day_workers = 0
        day_hours = 0
        day_amount = 0

        for request in day_requests:
            items = request['items']
            workers = request['workers']
            customer_amount = sum(worker['amount'] for worker in workers)
            worker_hours = sum(
                worker['hours']
                for worker in workers
                if worker['amount'] > 0
            )
            if worker_hours == 0:
                continue
            current_row = _write_interval_req_row(
                ws,
                current_row,
                [
                    request['date'],
                    request['route'],
                    [item['code'] for item in request['items']],
                    [item['address'] for item in request['items']],
                    '{} - {}'.format(
                        min(item['interval_begin'] for item in items)[:5],
                        max(item['interval_end'] for item in items)[:5],
                    ),
                    request['worker_turnout_count'],
                    worker_hours,
                    [worker['full_name'] for worker in workers],
                    [worker['hours'] for worker in workers],
                    [
                        worker['amount'] / worker['hours']
                        for worker in workers
                    ],
                    [worker['amount'] for worker in workers],
                    customer_amount,
                    request['worker_turnout_count'],
                    worker_hours,
                    request['confirmed_timepoint'],
                    (
                        timezone.localtime(request['arrival_time']).time().strftime('%H:%M')
                        if request['arrival_time'] is not None else '-'
                    ),
                    request['driver_name'],
                    request['driver_phones'],
                ],
                alignment=alignment_center,
                border=border_thin,
                font=font_calibri_10,
                fill=cell_fill_lightblue if len(request['items']) > 1 else None,
                number_formatting={
                    1: 'DD.MM.YY',
                    9: '0.00',
                    10: '0.00',
                    15: 'HH:mm',
                    16: 'HH:mm',
                }
            )

            day_workers += request['worker_turnout_count']
            day_hours += worker_hours
            day_amount += customer_amount

        current_row = _write_interval_req_row(
            ws,
            current_row,
            [
                day,
                '',
                '',
                '',
                '',
                day_workers,
                day_hours,
                '',
                '',
                '',
                day_amount,
                '',
                day_workers,
                day_hours,
                '',
                '',
                '',
                '',
            ],
            alignment=alignment_center,
            border=border_thin,
            font=font_calibri_10,
            fill=cell_fill_lightgreen,
            number_formatting={
                1: 'DD.MM.YY',
            }
        )

        total_workers += day_workers
        total_hours += day_hours
        total_amount += day_amount

        ws.merged_cells.ranges.append(tuple_to_range(
            current_row - 1,
            11,
            current_row - 1,
            12,
        ))

    current_row = _write_interval_req_row(
        ws,
        current_row,
        [
            'Итого',
            '',
            '',
            '',
            '',
            total_workers,
            total_hours,
            '',
            '',
            '',
            total_amount,
            '',
            total_workers,
            total_hours,
            '',
            '',
            '',
            '',
        ],
        alignment=alignment_center,
        border=border_thin,
        font=font_calibri_bold_10,
    )
    ws.merged_cells.ranges.append(tuple_to_range(
        current_row - 1,
        11,
        current_row - 1,
        12,
    ))

    ws.append([])
    current_row += 1

    ws.append(
        [openpyxl.cell.WriteOnlyCell(
            ws,
            value=(
                'Представитель Заказчика: __________________/______________________\n'
                '\n'
                'Исполнитель	\n'
                '\n'
                '____________________/____________________\n'
                'М.П.\n'
            )
        )] +
        [None] * 8 +
        [openpyxl.cell.WriteOnlyCell(
            ws,
            value=(
                'Заказчик\n'
                '\n'
                '____________________/____________________\n'
                'М.П.\n'
            )
        )],
    )
    ws.merged_cells.ranges.extend((
        tuple_to_range(current_row, 1, current_row, 8),
        tuple_to_range(current_row, 10, current_row, 16),
    ))
    current_row += 1


def _get_interval_report_queryset(first_day, last_day, customer_id, location_id):
    request_qs = DeliveryRequest.objects.filter(
        date__range=(first_day, last_day),
        status__in=DeliveryRequest.SUCCESS_STATUSES,
        customer=customer_id
    )
    no_turnout_sq = RequestWorker.objects.annotate(
        active=Exists(
            ItemWorker.objects.filter(
                requestworker=OuterRef('id'),
                itemworkerrejection__isnull=True,
            )
        )
    ).filter(
        active=True,
        workerrejection__isnull=True,
        request=OuterRef('id'),
        requestworkerturnout__isnull=True,
    )

    bad_requests = list(request_qs.annotate(
        no_turnouts=Exists(
            no_turnout_sq
        )
    ).filter(
        no_turnouts=True
    ).values_list('pk', flat=True))

    if bad_requests:  # todo filter by location before checking for turnouts?
        raise DataIntegrityException(
            f'В закрытых заявках {bad_requests} присутствуют работники без выходов.'
        )

    request_qs = request_qs.with_arrival_time(
    ).with_worker_count(
        turnout=True,
    )

    if location_id is not None:
        request_qs = request_qs.filter(location=location_id)

    return request_qs.with_confirmed_timepoint(
    ).with_items_count(
    ).with_items_for_customer(
        item_field_names=[
            'code',
            'address',
            'interval_begin',
            'interval_end',
        ]
    ).annotate(
        workers=Coalesce(
            Subquery(
                _get_worker_subquery().values(
                    'request'
                ).annotate(
                    workers_list=ArrayAgg(
                        'worker_fields',
                        ordering=(
                            'worker__last_name',
                            'worker__name',
                            'worker__patronymic',
                            'worker__id',
                        )
                    )
                ).values(
                    'workers_list'
                )
            ),
            [],
            output_field=ArrayField(JSONField())
        ),
    ).filter(
        worker_turnout_count__gt=0,
    ).order_by(
        'date',
        'items_count',
        'driver_name',
        'pk'
    ).values(
        'date',
        'route',
        'items',
        'arrival_time',
        'worker_turnout_count',
        'workers',
        'confirmed_timepoint',
        'driver_name',
        'driver_phones',
    )


def interval_report_response(first_day, last_day, customer, location):
    if location is None:
        location_id = None
        location_name = 'Все филиалы'
    else:
        location_id = location.pk
        location_name = location.location_name

    request_qs = _get_interval_report_queryset(
        first_day,
        last_day,
        customer.pk,
        location_id,
    )

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet()

    ws.page_setup.paperSize = openpyxl.worksheet.worksheet.Worksheet.PAPERSIZE_A4
    ws.page_setup.orientation = openpyxl.worksheet.worksheet.Worksheet.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    _fill_interval_sheet(ws, request_qs, first_day, last_day)

    return xlsx_content_response(
        wb,
        '{}_{}_{}-{}.xlsx'.format(
            customer.cust_name,
            location_name,
            string_from_date(first_day),
            string_from_date(last_day)
        )
    )


def suspicious_turnouts_report_response(first_day, last_day):
    def _arrivals_qs(is_suspicious):
        return ItemWorkerStart.objects.filter(
            itemworker__requestworker__request__date__range=(first_day, last_day)
        ).filter(
            is_suspicious=is_suspicious
        ).values(
            'itemworker__requestworker__worker',
            'itemworker__requestworker__request__date',
        ).annotate(
            Count('pk')
        ).order_by(
            'itemworker__requestworker__worker',
        ).values_list(
            'itemworker__requestworker__worker',
            'itemworker__requestworker__request__date',
            'pk__count',
        )

    def _arrivals(arrivals_qs):
        return {
            worker: {day: count for _, day, count in arrivals}
            for worker, arrivals in itertools.groupby(
                arrivals_qs,
                key=operator.itemgetter(0)
            )
        }

    suspicious_arrivals = _arrivals(_arrivals_qs(True))
    normal_arrivals = _arrivals(_arrivals_qs(False))

    workers = Worker.objects.with_full_name(
    ).filter(
        pk__in=suspicious_arrivals
    ).select_related(
        'workerzone__zone__name'
    ).order_by(
        'full_name',
    ).values_list(
        'pk',
        'full_name',
        'workerzone__zone__name'
    )

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet()

    ws.page_setup.paperSize = openpyxl.worksheet.worksheet.Worksheet.PAPERSIZE_A4
    ws.page_setup.orientation = openpyxl.worksheet.worksheet.Worksheet.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    days = days_from_interval(first_day, last_day)

    ws.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 30
    ws.column_dimensions[openpyxl.utils.get_column_letter(2)].width = 20
    for col_num in range(3, len(days) + 3):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 5

    font = font_calibri_bold_10
    cell_alignment = alignment_bottom_center

    def _make_cell(value, alignment=cell_alignment):
        cell = openpyxl.cell.WriteOnlyCell(ws, value=value)
        cell.alignment = alignment
        cell.font = font
        return cell

    ws.append(
        [_make_cell('ФИО')] +
        [_make_cell('Город')] +
        [_make_cell(day.strftime('%d.%m')) for day in days]
    )

    font = font_calibri_10

    for worker_id, worker_name, worker_zone in workers:
        _suspicious_arrivals = suspicious_arrivals[worker_id]
        _normal_arrivals = normal_arrivals.get(worker_id, {})
        def _get_suspicios_turnouts_count(day):
            count = _suspicious_arrivals.get(day)
            if count is None and day in _normal_arrivals:
                # Продуктовое желание: если нет подозрительных выходов, но
                # есть нормальные - писать в отчете '0'
                count = 0

            return count

        ws.append(
            [_make_cell(worker_name, alignment=alignment_bottom_left)] +
            [_make_cell(worker_zone, alignment=alignment_bottom_left)] +
            [_make_cell(_get_suspicios_turnouts_count(day)) for day in days]
        )

    return xlsx_content_response(
        wb,
        'Подозрительные_выходы_{}-{}.xlsx'.format(
            string_from_date(first_day),
            string_from_date(last_day)
        )
    )


# ### Online reports

# https://yandex.ru/dev/yandex-apps-launch/maps/doc/concepts/yandexmaps-web-docpage/
def _worker_locations(worker, day):
    mobile_locations = models.MobileAppStatus.objects.filter(
        timestamp__date=day,
        user=worker.workeruser.user
    ).filter(
        location__isnull=False
    ).order_by(
        'location__timestamp'
    ).values(
        'location__longitude',
        'location__latitude',
        'location__timestamp'
    ).iterator()

    def _filter(locs):
        try:
            prev = next(locs)
        except StopIteration:
            return

        yield prev

        while True:
            try:
                loc = next(locs)
            except StopIteration:
                return

            if (
                    abs(loc['location__longitude'] - prev['location__longitude']) > 0.00005 or
                    abs(loc['location__latitude'] - prev['location__latitude']) > 0.00005
            ):
                prev = loc
                yield loc

    def _serialize(location, idx):
        return '{:.6f},{:.6f},{}'.format(
            location['location__longitude'],
            location['location__latitude'],
            '{} ({})'.format(
                as_default_timezone(location['location__timestamp']).strftime('%H:%M'),
                idx
            )
        )

    return '~'.join([
        _serialize(location, i)
        for i, location in enumerate(_filter(mobile_locations))
    ])


def _replace_delimiters(text):
    replacements = (
        (',', '_'),
        (';', '_'),
        ('\r', ' '),
        ('\n', ' ')
    )
    for src, dst in replacements:
        text = text.replace(src, dst)
    return text


def _latlon(lat, lon):
    if lat is None or lon is None:
        lat = 55.751436 + ((random.random() - 0.5) / 2000)
        lon = 37.618886 + ((random.random() - 0.5) / 2000)
    return lat, lon


MAP_DATETIME_FORMAT = '%d.%m %H:%M'


def _worker_on_map_items_arrivals(rw):
    links = []
    operation_id = rw['requestworkerturnout__workerturnout__turnoutoperationtopay__operation']
    if operation_id is not None:
        links.append(  # amount link
            'сумма${}'.format(
                reverse(
                    'admin:finance_operation_change',
                    args=[operation_id]
                )
            )
        )
    links.append(  # photo link
        'фото$/rf/delivery/photos_dashboard/?request={}&worker={}'.format(
            rw['request'],
            rw['worker'],
        )
    )

    request_create_time = timezone.localtime(
        rw['request__timestamp']
    ).strftime(MAP_DATETIME_FORMAT)

    prefix = str(rw['request'])
    if len(rw['items']) > 1:
        prefix = 'M' + prefix
    if rw['request__status'] == DeliveryRequest.CANCELLED_WITH_PAYMENT:
        prefix = 'Х_П ' + prefix

    arrivals = []
    for arrival in rw['arrivals']:
        lat, lon = _latlon(arrival['latitude'], arrival['longitude'])
        arrivals.append(
            ','.join([
                f'{lon:.6f}',
                f'{lat:.6f}',
                _replace_delimiters(
                    '{} заявка {} фото {} {}/{}'.format(
                        prefix,
                        request_create_time,
                        timezone.localtime(
                            postgres_str_to_datetime(arrival['timestamp'])
                        ).strftime(MAP_DATETIME_FORMAT),
                        arrival['item_code'],
                        arrival['item_address'],
                    )
                ),
                *links
            ])
        )

    items = []
    route_prefix = 'M' if len(rw['items']) > 1 else ''
    for item in rw['items']:
        if item['geotag'] is not None:
            lat, lon = _latlon(item['geotag']['latitude'], item['geotag']['longitude'])
        else:
            lat, lon = _latlon(None, None)
        title = '{}{}: {}'.format(
            route_prefix,
            rw['request'],
            _replace_delimiters(item['address'])
        )
        items.append(
            ','.join([
                f'{lon:.6f}',
                f'{lat:.6f}',
                title,
            ])
        )

    return arrivals, items


def worker_on_map_link(worker, day):
    # Todo: '?ll=37.723778,55.718753&z=10'
    worker_on_map = RequestWorker.objects.filter(
        worker=worker,
        request__date=day,
        workerrejection__isnull=True,
    ).annotate(
        items=Coalesce(
            Subquery(
                DeliveryItem.objects.filter(
                    request=OuterRef('request')
                ).with_geotag(
                ).annotate(
                    item_fields=JSONObject(
                        code='code',
                        address='address',
                        geotag='geotag_',
                    )
                ).order_by(
                ).values(
                    'request'
                ).annotate(
                    items_list=ArrayAgg(
                        'item_fields',
                        ordering=(
                            'pk',
                        )
                    )
                ).values(
                    'items_list'
                ),
                output_field=ArrayField(JSONField())
            ),
            []
        ),
        arrivals=Coalesce(
            Subquery(
                ItemWorkerStart.objects.filter(
                    itemworker__requestworker=OuterRef('pk'),
                    itemworker__itemworkerrejection__isnull=True,
                ).annotate(
                    arrival_fields=JSONObject(
                        item_code='itemworker__item__code',
                        item_address='itemworker__item__address',
                        latitude='location__latitude',
                        longitude='location__longitude',
                        timestamp='timestamp',
                    )
                ).values(
                    'itemworker__requestworker'
                ).annotate(
                    arrivals_list=ArrayAgg(
                        'arrival_fields',
                        ordering=(
                            'timestamp',
                        )
                    )
                ).values(
                    'arrivals_list'
                ),
                output_field=ArrayField(JSONField())
            ),
            []
        )
    ).values(
        'request',
        'worker',
        'request__status',
        'request__timestamp',
        'requestworkerturnout__workerturnout__turnoutoperationtopay__operation',
        'items',
        'arrivals',
    ).order_by(
        'request__timestamp'
    )

    try:
        arrivals, items = zip(*(_worker_on_map_items_arrivals(rw) for rw in worker_on_map))
    except ValueError:
        arrivals = items = []

    return {
        'title': '{} {}'.format(day.strftime('%d.%m'), str(worker)),
        'pt': _worker_locations(worker, day),
        'pta': '~'.join(arrival for req in arrivals for arrival in req),
        'ptr': '~'.join(item for req in items for item in req),
    }


def _request_description(worker_request):
    start_time = as_default_timezone(worker_request.start).strftime('%H:%M')
    if worker_request.finish is not None:
        finish_time = as_default_timezone(worker_request.finish).strftime('%H:%M')
        duration = time_interval_format(
            worker_request.finish - worker_request.start
        )
    else:
        finish_time = '???'
        duration = '???'
    return '\n'.join(
        [
            f'{start_time} - {finish_time} ({duration})',
            f'{worker_request.request_id}, {worker_request.operator}',
            f'{sum(item["mass"] for item in worker_request.items)} кг',
        ] + [item["address"] for item in worker_request.items]
    )


def _request_count_texts(count):
    r100 = count % 100
    r10 = count % 10
    if r100 == 1 or (r100 > 20 and r10 == 1):
        verb = 'осталась'
        name = 'заявка'
    elif r100 in [2, 3, 4] or (r100 > 20 and r10 in [2, 3, 4]):
        verb = 'осталось'
        name = 'заявки'
    else:
        verb = 'осталось'
        name = 'заявок'

    return verb, name


def turnouts_report(request):
    day = _get_value(request, 'date')
    if day:
        day = date_from_string(day)
    else:
        day = timezone.now().date()

    day_range = (
        timezone.make_aware(
            datetime.datetime.combine(
                day,
                datetime.time(hour=4),
            )
        ),
        timezone.make_aware(
            datetime.datetime.combine(
                day + datetime.timedelta(days=1),
                datetime.time(hour=4),
            )
        )
    )

    worker_requests = list(
        RequestWorker.objects.filter(
            request__date=day,
            workerrejection__isnull=True,
        ).annotate(
            has_active_items=Exists(
                ItemWorker.objects.filter(
                    requestworker=OuterRef('pk'),
                    itemworkerrejection__isnull=True,
                )
            ),
            start=Subquery(
                ItemWorkerStart.objects.filter(
                    itemworker__requestworker=OuterRef('pk'),
                    itemworker__itemworkerrejection__isnull=True,
                ).order_by(
                    'timestamp'
                ).values(
                    'timestamp'
                )[:1],
                output_field=DateTimeField()
            ),
            finish=Subquery(
                ItemWorkerFinish.objects.filter(
                    itemworker__requestworker=OuterRef('pk'),
                    itemworker__itemworkerrejection__isnull=True,
                ).order_by(
                    '-timestamp'
                ).values(
                    'timestamp'
                )[:1],
                output_field=DateTimeField()
            ),
            amount=Subquery(
                RequestWorkerTurnout.objects.filter(
                    requestworker=OuterRef('pk'),
                ).values(
                    'workerturnout__turnoutoperationtopay__operation__amount'
                ),
                output_field=DecimalField()
            ),
            operator=Subquery(
                User.objects.filter(
                    deliveryrequestoperator__request=OuterRef('request')
                ).values('first_name')
            ),
            items=Subquery(
                ItemWorker.objects.filter(
                    requestworker=OuterRef('pk'),
                    itemworkerrejection__isnull=True,
                ).annotate(
                    item_fields=JSONObject(
                        mass='item__mass',
                        address='item__address',
                    )
                ).order_by(
                ).values(
                    'requestworker'
                ).annotate(
                    items_list=ArrayAgg(
                        'item_fields',
                        ordering=(
                            'itemworkerstart',
                            'item__interval_begin',
                            'item__interval_end',
                            'item_id',
                        )
                    ),
                ).values_list(
                    'items_list'
                ),
                output_field=ArrayField(JSONField())
            ),
        ).filter(
            has_active_items=True,
            start__isnull=False
        ).order_by(
            'worker',
            'start'
        )
    )

    requests_for_worker = {
        worker_id: list(w_requests)
        for worker_id, w_requests in itertools.groupby(
            worker_requests,
            key=operator.attrgetter('worker_id')
        )
    }

    workers = Worker.objects.filter(
        pk__in=requests_for_worker
    ).with_age(
    ).annotate(
        not_newbie=Exists(
            WorkerTurnout.objects.filter(
                worker=OuterRef('pk'),
                timesheet__sheet_date__lt=day
            )
        ),
        locations=Subquery(
            MobileAppStatus.objects.filter(
                timestamp__range=day_range,
                user=OuterRef('workeruser__user'),
                location__isnull=False,
            ).annotate(
                status_fields=JSONObject(
                    timestamp='location__timestamp',
                    latitude='location__latitude',
                    longitude='location__longitude',
                )
            ).values(
                'user'
            ).annotate(
                status_list=ArrayAgg(
                    'status_fields',
                    ordering=(
                        'timestamp',
                    ),
                ),
            ).values(
                'status_list'
            ),
            output_field=ArrayField(JSONField())
        ),
        online=Subquery(
            MobileAppStatus.objects.filter(
                timestamp__range=day_range,
                user=OuterRef('workeruser__user')
            ).values(
                'user'
            ).annotate(
                Min('timestamp'),
                Max('timestamp'),
                Max('version_code'),
            ).annotate(
                status_fields=JSONObject(
                    start='timestamp__min',
                    finish='timestamp__max',
                    version='version_code__max',
                )
            ).values(
                'status_fields'
            ),
            output_field=JSONField()
        ),
        zone_name=Coalesce(
            F('workerzone__zone__name'),
            Value('-')
        )
    ).select_related(
        'citizenship'
    ).order_by(
        'workerzone__zone__name',
        'pk',
    )

    max_requests = max(map(len, requests_for_worker.values()), default=0)

    for worker in workers:

        requests = requests_for_worker[worker.pk]
        if worker.online is not None:
            worker.online['start'] = timezone.localtime(
                postgres_str_to_datetime(worker.online['start'])
            )
            worker.online['finish'] = timezone.localtime(
                postgres_str_to_datetime(worker.online['finish'])
            )

        worker.slots = slots_chain([
            (r.start, r.finish, r)
            for r in requests
        ])
        for slot in worker.slots:
            slot.length = slot.finish_index - slot.start_index + 1
            slot.empty_cells = range(slot.length - 1)
            if slot.objects:
                amount = int(sum([r.amount for r in slot.objects if r.amount is not None]))
                items_num = sum([len(r.items) for r in slot.objects])
                hour_rate = round(amount * 2 / slot.length)
                slot.text = f'{amount} ({hour_rate}/ч)/{items_num}'
                requests_count = len(slot.objects)
                if requests_count > 1:
                    _, name = _request_count_texts(requests_count)
                    slot.text += f' ({requests_count} {name})'
                slot.title = '\n'.join([_request_description(r) for r in slot.objects])

        worker.day_mass = 0.
        worker.day_amount = ZERO_OO
        worker.day_requests = len(requests)
        for r in requests:
            worker.day_mass += sum(item['mass'] for item in r.items)
            if r.amount:
                worker.day_amount += r.amount

        worker.day_mass = round(worker.day_mass, 1)

    return render(
        request,
        'the_redhuman_is/delivery/turnouts_report.html',
        {
            'day': day,
            'form': forms.SingleDateForm(initial={'date': day}),
            'slots': [time_interval_format(slot_start) for slot_start in slot_starts],
            'workers': workers,
            'requests_indexes': range(max_requests)
        }
    )


def _serialize_itemworkerstart(itemworkerstart):
    return {
        'request': {
            'id': itemworkerstart['itemworker__requestworker__request'],
            'date': itemworkerstart['itemworker__requestworker__request__date'],
            'driver_name': itemworkerstart['itemworker__requestworker__request__driver_name'],
            'operator': itemworkerstart['operator_name'],
            'location': itemworkerstart[
                'itemworker__requestworker__request__location__location_name'
            ],
            'status': itemworkerstart['itemworker__requestworker__request__status_description'],
            'comment': itemworkerstart['itemworker__requestworker__request__comment'],
            'extra_photo_count': itemworkerstart['extra_photo_count'],
        },
        'item': {
            'id': itemworkerstart['itemworker__item'],
            'code': itemworkerstart['itemworker__item__code'],
            'address': itemworkerstart['itemworker__item__address'],
        },
        'worker': {
            'id': itemworkerstart['itemworker__requestworker__worker'],
            'name': itemworkerstart['worker_name'],
        },
        'discrepancy': {
            'suspicious': itemworkerstart['is_suspicious'],
            'is_ok': itemworkerstart['itemworkerdiscrepancycheck__is_ok'],
            'comment': itemworkerstart['itemworkerdiscrepancycheck__comment'] or '',
        }
    }


def turnouts_period_report(request):
    first_day, last_day = _get_first_last_day(request)
    try:
        filter_distance = Decimal(_get_value(request, 'filter_distance'))
    except (TypeError, ValueError):
        filter_distance = Decimal('300')

    checked_status = _get_value(request, 'checked_status', 'all')

    address_sq = NormalizedAddress.objects.filter(
        location=OuterRef('itemworker__item'),
        version=OuterRef('itemworker__item__address_version'),
    )

    itemworkerstart_qs = ItemWorkerStart.objects.filter(
        itemworker__requestworker__request__date__range=(first_day, last_day),
    ).annotate(
        latitude=Subquery(
            address_sq.values('latitude')
        ),
        longitude=Subquery(
            address_sq.values('longitude')
        ),
    ).annotate(
        worker_name=PostgresConcatWS(
            Value(' '),
            F('itemworker__requestworker__worker__last_name'),
            F('itemworker__requestworker__worker__name'),
            NullIf(F('itemworker__requestworker__worker__patronymic'), Value(''))
        ),
        operator_name=Coalesce(
            NullIf(
                'itemworker__requestworker__request__deliveryrequestoperator__operator__first_name',
                Value('')
            ),
            'itemworker__requestworker__request__deliveryrequestoperator__operator__username',
        ),
        extra_photo_count=Coalesce(
            Subquery(
                Photo.objects.filter(
                    content_type=ContentType.objects.get_for_model(DeliveryRequest),
                    object_id=OuterRef('itemworker__requestworker__request'),
                ).values(
                    'object_id'
                ).annotate(
                    Count('pk')
                ).values(
                    'pk__count'
                )
            ),
            0,
            output_field=IntegerField()
        ),
        distance=Haversine(
            'latitude',
            'location__latitude',
            'longitude',
            'location__longitude',
            output_field=DecimalField(),
        )
    ).filter(
        Q(distance__isnull=True) |
        Q(distance__gt=filter_distance / 1000),
        latitude__isnull=False,
        longitude__isnull=False,
    ).values(
        'itemworker__requestworker__worker',
        'worker_name',
        'itemworker__requestworker__request',
        'itemworker__requestworker__request__date',
        'operator_name',
        'itemworker__requestworker__request__driver_name',
        'itemworker__requestworker__request__location__location_name',
        'itemworker__requestworker__request__status_description',
        'itemworker__item',
        'itemworker__item__code',
        'itemworker__item__address',
        'itemworker__requestworker__request__comment',
        'itemworkerdiscrepancycheck__is_ok',
        'is_suspicious',
        'itemworkerdiscrepancycheck__comment',
        'extra_photo_count',
        'distance',
    ).order_by(
        'timestamp'
    )

    if checked_status == 'not_checked':
        itemworkerstart_qs = itemworkerstart_qs.filter(
            Q(itemworkerdiscrepancycheck__isnull=True) |
            Q(itemworkerdiscrepancycheck__is_ok=None)
        )
    elif checked_status == 'ok':
        itemworkerstart_qs = itemworkerstart_qs.filter(itemworkerdiscrepancycheck__is_ok=True)
    elif checked_status == 'suspicious':
        itemworkerstart_qs = itemworkerstart_qs.filter(itemworkerdiscrepancycheck__is_ok=False)

    arrivals = list(
        _serialize_itemworkerstart(itemworkerstart) for itemworkerstart in itemworkerstart_qs
    )

    return render(
        request,
        'the_redhuman_is/delivery/turnouts_period_report.html',
        {
            'first_day': first_day,
            'last_day': last_day,
            'filter_distance': filter_distance,
            'arrivals': arrivals,
        }
    )


@csrf_exempt
def geo_map(request):
    points = []

    min_lat = 90.0
    max_lat = -90.0
    min_lon = 180.0
    max_lon = -180.0

    def _decode(pts):
        result = []
        if pts:
            for point in pts.split('~'):
                info = point.split(',')
                lon, lat, title = tuple(info[:3])

                nonlocal min_lat
                nonlocal max_lat
                nonlocal min_lon
                nonlocal max_lon

                f_lon = float(lon)
                f_lat = float(lat)

                if f_lon > max_lon:
                    max_lon = f_lon
                if f_lon < min_lon:
                    min_lon = f_lon
                if f_lat > max_lat:
                    max_lat = f_lat
                if f_lat < min_lat:
                    min_lat = f_lat

                urls = []
                for link in info[3:]:
                    label, url = link.split('$')
                    urls.append((label, url))
                result.append((lon, lat, title, urls))
        return result

    def _add_points(key, preset):
        points.append(
            {
                'preset': preset,
                'pts': _decode(_get_value(request, key))
            }
        )

    _add_points('pt', 'islands#blueCircleDotIcon')
    _add_points('pta', 'islands#greenDotIcon')
    _add_points('ptr', 'islands#redDotIcon')

    distance = geo_utils.distance(
        min_lon,
        min_lat,
        max_lon,
        max_lat
    )

    zoom = 13
    if distance > 3000000:
        zoom = 3
    elif distance > 400000:
        zoom = 5
    elif distance > 25000:
        zoom = 11
    elif distance > 15000:
        zoom = 12

    return render(
        request,
        'the_redhuman_is/delivery/map.html',
        {
            'title': _get_value(request, 'title'),
            'center_lat': str((min_lat + max_lat) / 2.0),
            'center_lon': str((min_lon + max_lon) / 2.0),
            'points': points,
            'zoom': zoom
        }
    )


# ### Statistics


def requests_count_report(request):
    return redirect('/rf/delivery/requests_count/')


def requests_count_report_data(request):
    first_day, last_day = _get_first_last_day(request)
    with_score = strtobool(_get_value(request, 'score'))
    with_score_v2 = strtobool(_get_value(request, 'score_v2'))
    with_score_v3 = strtobool(_get_value(request, 'score_v3'))
    no_locations = strtobool(_get_value(request, 'no_locations', 'false'))
    operator_pk = _get_value(request, 'operator')
    customer_pk = _get_value(request, 'customer')

    statuses = DeliveryRequest.SUCCESS_STATUSES + [
        DeliveryRequest.CANCELLED,
        DeliveryRequest.FAILED,
    ]

    delivery_request_qs = DeliveryRequest.objects.filter(
        date__range=(first_day, last_day),
        status__in=statuses,
    )
    if operator_pk:
        delivery_request_qs = delivery_request_qs.filter(
            deliveryrequestoperator__operator_id=operator_pk
        )
    if customer_pk:
        delivery_request_qs = delivery_request_qs.filter(
            customer__pk=customer_pk
        )

    # 1 собираем "объекты" для подсчета - Customer целиком или CustomerLocation
    if no_locations:
        objects_to_annotate = models.Customer.objects.annotate(
            has_requests=Exists(
                delivery_request_qs.filter(customer=OuterRef('pk'))
            ),
        )
    else:
        objects_to_annotate = models.CustomerLocation.objects.annotate(
            has_requests=Exists(
                delivery_request_qs.filter(location=OuterRef('pk'))
            ),
            cust_name=F('customer_id__cust_name')
        )
    objects_to_annotate = objects_to_annotate.filter(
        has_requests=True
    )

    # 2 добавляем по каждому доп. аннотации, если можем (для случая scores_v2 - не можем
    # посчитать через аннотации)
    count_overdue = (
        not with_score and
        not with_score_v2 and
        not with_score_v3 and
        no_locations
    )
    values_to_count = statuses + (['overdue'] if count_overdue else [])

    def _with_score(status, request_qs, force_no_locations=no_locations):
        if with_score or with_score_v3:
            if force_no_locations:
                requests_to_count = request_qs.filter(
                    customer=OuterRef('pk'),
                    status=status,
                )
            else:
                requests_to_count = request_qs.filter(
                    location=OuterRef('pk'),
                    status=status,
                )

            if with_score:
                requests_to_count = requests_to_count.with_score()
            else:
                requests_to_count = requests_to_count.with_score_v2()

            if force_no_locations:
                requests_to_count = requests_to_count.values(
                    'customer_id',
                )
            else:
                requests_to_count = requests_to_count.values(
                    'location_id',
                )

            return Coalesce(
                Subquery(
                    requests_to_count.annotate(
                        sum=Sum('score'),
                    ).values_list(
                        'sum',
                    ),
                    output_field=IntegerField(),
                ),
                0
            )
        else:
            return Count(
                'deliveryrequest',
                filter=Q(
                    deliveryrequest__in=request_qs,
                    deliveryrequest__status=status,
                )
            )

    values = ['pk', 'cust_name']
    if not no_locations:
        values.extend(('location_name', 'customer_id'))

    if not with_score_v2:
        objects_to_annotate = objects_to_annotate.annotate(
            **{
                status: _with_score(status, delivery_request_qs) for status in statuses
            }
        ).annotate(
            total=reduce(add, (F(status) for status in statuses)),
        )

        if count_overdue:
            objects_to_annotate = objects_to_annotate.annotate(
                overdue=Count(
                    'deliveryrequest',
                    filter=Q(
                        deliveryrequest__in=delivery_request_qs.filter(
                            status__in=DeliveryRequest.SUCCESS_STATUSES,
                        ).annotate(
                            last_confirmation=Subquery(
                                RequestWorkerTurnout.objects.filter(
                                    requestworker__request=OuterRef('pk'),
                                ).order_by(
                                    '-timestamp'
                                ).values('timestamp')[:1]
                            ),
                            duration=F('last_confirmation') - F('date')
                        ).filter(
                            duration__gt=Value(datetime.timedelta(minutes=29*60))
                        )
                    )
                )
            )

        objects_to_annotate = objects_to_annotate.filter(
            total__gt=0
        )

        values.extend(values_to_count)

    # 3 сортируем, преобразуем в список словариков
    if no_locations:
        objects_to_annotate = objects_to_annotate.order_by('cust_name')
    else:
        objects_to_annotate = objects_to_annotate.order_by('cust_name', 'location_name')

    objects_to_annotate = objects_to_annotate.values(
        *values
    )

    annotated_list = list(objects_to_annotate)

    # 4 для "сложных очков" считаем в питоне
    def _count_score(request_qs):
        requests = list(
            request_qs.with_score_v2(
            ).with_merge_data(
            )
        )

        for request in requests:
            if request.requests_to_merge_with:
                if request.pk < request.requests_to_merge_with[0]:
                    request.score = RequestWorkerTurnout.objects.filter(
                        requestworker__request__in=request.requests_to_merge_with + [request.pk]
                    ).values(
                        'requestworker__worker'
                    ).annotate(
                        score=Ceil(
                            Sum('workerturnout__hours_worked') /
                            Value(6, output_field=IntegerField()),
                            output_field=IntegerField()
                        )
                    ).aggregate(
                        Sum('score')
                    )['score__sum']
                else:
                    request.score = 0

        return sum([request.score for request in requests])

    if with_score_v2:
        tmp_annotated_list = annotated_list
        annotated_list = []
        for item in tmp_annotated_list:
            total = 0
            for status in statuses:
                if no_locations:
                    item_requests_qs = delivery_request_qs.filter(
                        customer=item['pk'],
                        status=status,
                    )
                else:
                    item_requests_qs = delivery_request_qs.filter(
                        location=item['pk'],
                        status=status,
                    )
                score = _count_score(item_requests_qs)
                total += score
                item[status] = score
            item['total'] = total
            if total > 0:
                annotated_list.append(item)

    data = {}
    prev_customer = None
    for item in annotated_list:
        if no_locations:
            label = item['cust_name']
        else:
            # Try to add requests without location
            current_customer = item['customer_id']
            if prev_customer is None or prev_customer != current_customer:
                prev_customer = current_customer
                requests_without_location = delivery_request_qs.filter(
                    customer=prev_customer,
                    location__isnull=True,
                    status__in=statuses
                )
                if requests_without_location.exists():
                    customer = models.Customer.objects.filter(
                        pk=prev_customer
                    )
                    if with_score_v2:
                        customer = customer.values('pk', 'cust_name')
                        customer = customer.get()

                        total = 0
                        for status in statuses:
                            score = _count_score(
                                requests_without_location.filter(status=status)
                            )
                            total += score
                            customer[status] = score
                        customer['total'] = total
                    else:
                        customer = customer.annotate(
                            **{
                                status: _with_score(status, requests_without_location, True)
                                for status in statuses
                            }
                        ).annotate(
                            total=reduce(add, (F(status) for status in statuses)),
                        ).values(
                            *(values_to_count + ['cust_name'])
                        )
                        customer = customer.get()

                    data[f'{customer["cust_name"]}/Без объекта'] = {
                        f'{status}_count': customer[status]
                        for status in values_to_count
                    }

            label = item['cust_name'] + '/' + item['location_name']
        data[label] = {
            f'{status}_count': item[status] for status in values_to_count
        }

    def _status_total(status_name):
        return sum(e[f'{status_name}_count'] for e in data.values())
    data['Итого'] = {
        f'{status}_count': _status_total(status) for status in values_to_count
    }

    return JsonResponse({'status': 'ok', 'data': data})


def imports_report(request):
    return redirect('/rf/delivery/imports_report/')


def imports_report_data(request):
    first_day, last_day = _get_first_last_day(request)
    customer_pk = _get_value(request, 'customer')

    files = models.RequestsFile.objects.filter(
        timestamp__date__range=(first_day, last_day)
    ).select_related(
        'author'
    ).order_by('-pk')
    if customer_pk:
        files = files.filter(customer__pk=customer_pk)

    base_url = request.build_absolute_uri(reverse('the_redhuman_is:delivery_requests_file'))
    data = [_serialize_requests_file(f, base_url) for f in files]

    return JsonResponse(data, safe=False)


def requests_file(request):
    pk = _get_value(request, 'pk')
    file_type = _get_value(request, 'type')

    return requests_file_response(pk, file_type)


def cities_report(request):
    result = {}

    addresses = NormalizedAddress.objects.all()
    for address in addresses:
        data = json.loads(address.raw_data)
        if len(data) != 1:
            print('wtf? {}'.format(len(data)))
        data = data[0]
        region = data.get('region_with_type')
        if region not in result:
            result[region] = {}

        city = data.get('city')
        if not city:
            city = data.get('area')

        if not city:
            continue

        if city not in result[region]:
            result[region][city] = 1
        else:
            result[region][city] += 1

    wb = openpyxl.Workbook()
    ws = wb.active

    for region, cities in result.items():
        sorted_cities = [(city, count) for city, count in cities.items()]
        sorted_cities.sort(key=lambda v: -v[1])

        sheet = wb.create_sheet(str(region).replace('/', '_'))

        sheet.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 30

        row = 1
        for city, count in sorted_cities:
            sheet.cell(row=row, column=1, value=city)
            sheet.cell(row=row, column=2, value=count)
            row += 1

    wb.remove_sheet(ws)

    return xlsx_content_response(wb, 'cities.xlsx')


def new_customers_report(request):
    return render(
        request,
        'delivery_customers_page.html',
        {
        }
    )


def new_customers_report_data(request):
    legal_entities = models.DeliveryCustomerLegalEntity.objects.filter(
        registration_confirmed=False
    ).order_by('-pk')

    def _serialize(legal_entity):
        data = {}

        fields = [
            'pk',
            'full_name',
            'email',
            'phone',
            'tax_number',
            'reason_code',
            'bank_name',
            'legal_address',
            'mail_address',
        ]
        for field in fields:
            data[field] = getattr(legal_entity, field)

        if legal_entity.is_legal_entity is None:
            data['type'] = '-'
        else:
            data['type'] = 'ООО' if legal_entity.is_legal_entity else 'ИП'

        data['scans'] = []
        photos = models.get_photos(legal_entity)
        for photo in photos:
            data['scans'].append(request.build_absolute_uri(photo.image.url))

        return data

    return JsonResponse([_serialize(e) for e in legal_entities], safe=False)


# ### Workers, drivers


@require_GET
def other(request):
    call_form = forms.DeliverZoneForm(initial={'zone': DeliveryZone.objects.get(code='msk')})
    call_form.helper = FormHelper()
    call_form.helper.layout = Layout(
        Row(
            Field('zone', css_class='form-control form-control-sm'),
            SubmitNoValue(
                'submit',
                'Скачать',
                css_class='btn-sm btn-action btn-outline-primary'
            ),
            css_class='form-inline mb-0 form-row'
        )
    )
    call_form.helper.form_action = reverse('the_redhuman_is:delivery_call_list_csv')
    last_day = timezone.localdate()
    first_day = last_day - datetime.timedelta(days=30)
    return render(
        request,
        'the_redhuman_is/delivery/other.html',
        {
            'call_form': call_form,
            'interval_form': forms.DaysPickerIntervalForm(
                initial={
                    'first_day': first_day,
                    'last_day': last_day
                }
            ),
        }
    )


@require_POST
def worker_call_list_csv(request):
    form = forms.DeliverZoneForm(request.POST)
    if not form.is_valid():
        return redirect('the_redhuman_is:delivery_call_list')
    zone = form.cleaned_data['zone']

    workers = Worker.objects.all(
    ).with_full_name(
    ).filter_mobile(
    ).filter(
        workerzone__zone=zone.pk,
        banned__isnull=True,
    ).annotate(
        last_turnout_date=Subquery(
            WorkerTurnout.objects.filter(
                worker=OuterRef('pk')
            ).order_by(
                '-timesheet__sheet_date'
            ).values(
                'timesheet__sheet_date'
            )[:1]
        ),
    ).order_by(
        '-last_turnout_date',
        '-input_date'
    ).values(
        'last_turnout_date',
        'input_date',
        'full_name',
        'tel_number',
    )

    response = HttpResponse(content_type='text/csv', charset='windows-1251')
    response['Content-Disposition'] = "attachment; filename*=UTF-8''{}".format(
        quote(f'{zone.code}_workers.csv')
    )
    writer = csv.writer(response, delimiter=';')
    for worker in workers:
        writer.writerow((
            string_from_date(worker['last_turnout_date']) or '-',
            string_from_date(worker['input_date']),
            worker['full_name'],
            worker['tel_number'],
        ))
    return response


# Todo: this is a temporary view!
def online_status_report(request):
    date = date_from_string(_get_value(request, 'date'))
    zone = _get_value(request, 'zone')
    if date is None:
        date = timezone.localdate()
    show_rejected = strtobool(_get_value(request, 'show_rejected'))

    marks = OnlineStatusMark.objects.filter(
        timestamp__date=date - datetime.timedelta(days=1),
        user__workeruser__worker=OuterRef('pk')
    )

    workers = Worker.objects.all(
    ).select_related(
        'citizenship'
    ).annotate(
        has_online_mark=Exists(
            marks
        )
    ).filter(
        has_online_mark=True
    ).annotate(
        online=Subquery(
            marks.order_by(
                '-timestamp'
            ).values(
                'online'
            )[:1]
        ),
    ).annotate(
        zone_name=Subquery(
            ZoneGroup.objects.filter(
                workerzone__worker=OuterRef('pk')
            ).values(
                'name'
            )
        ),
        has_turnouts=Exists(
            WorkerTurnout.objects.filter(
                worker=OuterRef('pk')
            )
        ),
        requests_count=Coalesce(
            Subquery(
                RequestWorker.objects.annotate(
                    is_active=Exists(
                        ItemWorker.objects.filter(
                            requestworker=OuterRef('pk'),
                            itemworkerrejection__isnull=True
                        )
                    )
                ).filter(
                    request__date=date,
                    worker=OuterRef('pk'),
                    workerrejection__isnull=True,
                    is_active=True,
                ).values(
                    'worker'
                ).annotate(
                    Count('pk')
                ).values(
                    'pk__count'
                ),
                output_field=IntegerField()
            ),
            0
        )
    )

    if zone is not None:
        workers = workers.filter(
            workerzone__zone__pk=zone
        )

    if not show_rejected:
        workers = workers.filter(online=True)

    zone_form = forms.DeliverZoneForm(initial={'zone': zone})
    zone_form.fields['zone'].required = False
    return render(
        request,
        'the_redhuman_is/delivery/online_status_report.html',
        {
            'date': date,
            'date_form': forms.SingleDateForm(initial={'date': date}),
            'zone_form': zone_form,
            'show_rejected': not show_rejected,
            'workers': workers,
        }
    )


# This may be a temporary view
def workers_to_connect_to_mts(request):
    today = timezone.localdate()
    recent_request_threshold = today - datetime.timedelta(days=10)

    workers = Worker.objects.all(
    ).with_selfemployment_tax_number(
    ).annotate(
        has_recent_requests=Exists(
            RequestWorker.objects.filter(
                request__date__gte=recent_request_threshold,
                worker=OuterRef('pk'),
            )
        ),
        has_receipts=Exists(
            WorkerReceipt.objects.filter(
                worker=OuterRef('pk')
            )
        )
    ).filter(
        has_recent_requests=True,
        tax_number__isnull=False,
        has_receipts=False,
    ).values_list(
        'last_name',
        'name',
        'patronymic',
        'tax_number',
        'tel_number',
    )

    response = HttpResponse(content_type='text/csv', charset='windows-1251')
    response['Content-Disposition'] = "attachment; filename*=UTF-8''{}".format(
        quote(f'Запрос прав {string_from_date(today)}.csv')
    )

    writer = csv.writer(response, delimiter=';')
    for row in workers:
        writer.writerow(row + (settings.COMPANY_NAME,))

    return response


@user_passes_test(lambda user: user.is_superuser)
def confirm_legal_entity(request):
    legal_entity = models.DeliveryCustomerLegalEntity.objects.get(pk=_get_value(request, 'pk'))

    legal_entity.registration_confirmed = True
    legal_entity.save()

    return HttpResponse()


# Todo: new models (?), move to service functions, etc
def operators_for_driver(request):
    username = _get_value(request, 'login')
    password = _get_value(request, 'password')
    driver_phone = _get_value(request, 'driver_phone')

    user = authenticate(username=username, password=password)
    if user is None or driver_phone in ['', None]:
        return HttpResponseNotFound()

    user_groups = user.groups.values_list('name', flat=True)
    if 'Интеграция с телефонией' not in user_groups:
        return HttpResponseForbidden()

    now = as_default_timezone(timezone.now())
    today = now.date()

    driver_requests = DeliveryRequest.objects.filter(
        date=today,
        driver_phones__icontains=driver_phone,
    )

    driver_zones = ZoneGroup.objects.annotate(
        request_exists=Exists(
            driver_requests.filter(
                location__locationzonegroup__zone_group=OuterRef('pk')
            )
        )
    ).filter(
        request_exists=True
    )
    if not driver_zones.exists():
        driver_zones = ZoneGroup.objects.filter(code='msk')

    if now.time() < datetime.time(hour=17, minute=0):
        main_operators = User.objects.annotate(
            request_exists=Exists(
                driver_requests.filter(deliveryrequestoperator__operator=OuterRef('pk'))
            )
        ).filter(
            request_exists=True
        )
    else:
        main_operators = User.objects.annotate(
            operator_exists=Exists(
                OperatorZoneGroup.objects.filter(
                    operator=OuterRef('pk'),
                    zone_group__in=driver_zones
                )
            )
        ).filter(
            operator_exists=True
        )
    main_operators = list(main_operators.values_list('username', flat=True))

    driver_secondary_operators = User.objects.annotate(
        request_exists=Exists(
            DeliveryRequest.objects.filter(
                date=today,
                location__locationzonegroup__zone_group__in=driver_zones,
                deliveryrequestoperator__operator=OuterRef('pk')
            )
        )
    ).filter(
        request_exists=True
    ).values_list(
        'username',
        flat=True
    )

    secondary_operators = [
        username for username in driver_secondary_operators if username not in main_operators
    ]

    driver_names = set(
        driver_requests.values_list('driver_name', flat=True)
    )
    if len(driver_names) == 1:
        driver_name = list(driver_names)[0]
    else:
        driver_name = None

    return JsonResponse(
        {
            'driver_name': driver_name,
            'main_operators': ['delivery_42', 'delivery_61'], # main_operators,
            'secondary_operators': [], # secondary_operators
        }
    )


# ### Mobile API

MIN_ANDROID_APP_VERSION_CODE = 10046

# errors:
#  'banned',
#  'unauthenticated',
#  'outdated_version',
#  'expired_documents',
#  'is_not_selfemployed',
#  'suspicious_location',
#  'payout_already_requested',
def mobile_api(view_func):
    @api_view(['POST'])
    @transaction.atomic
    def proxy(request, *args, **kwargs):
        print('processing request')
        user = request.user
        if not user.is_authenticated:
            print('error: general unauthenticated')
            return HttpResponseForbidden(
                json.dumps({'error': 'unauthenticated'}),
                content_type='application/json'
            )

        status = request.POST.get('status')
        print('status: {}'.format(status))
        outdated = True
        if status:
            status = json.loads(status)

            version_code = status.get('version_code')
            if (version_code is not None) and (version_code >= MIN_ANDROID_APP_VERSION_CODE):
                outdated = False

                fcm_app_id = status.get('fcm_app_id')
                fcm_token = status.get('fcm_token')
                if fcm_app_id and fcm_token:
                    update_fcm_token(user, fcm_app_id, fcm_token)
                else:
                    print(f'Warning: there is no FCM token for user {user}')

                provider = status.get('location_provider')
                latitude = status.get('location_lat')
                longitude = status.get('location_lon')
                location_time = status.get('location_time')
                base_loc = None
                if provider and latitude and longitude and location_time:
                    base_loc = Location.objects.create(
                        provider=provider,
                        latitude=latitude,
                        longitude=longitude,
                        time=location_time
                    )
                setattr(request, 'location', base_loc)

                app_status = models.MobileAppStatus.objects.create(
                    user=user,
                    app_id=fcm_app_id,
                    version_code=version_code,
                    device_manufacturer=status.get('device_manufacturer'),
                    device_model=status.get('device_model'),
                    location=base_loc,
                )
                worker_id = WorkerUser.objects.values_list(
                    'worker_id',
                    flat=True,
                ).filter(
                    user_id=user.pk
                ).first()

                if worker_id is not None:
                    MobileAppWorker.objects.get_or_create(worker_id=worker_id)
                    WorkerRating.objects.get_or_create(worker_id=worker_id)

                    if base_loc is not None:
                        LastLocation.objects.update_or_create(
                            worker_id=worker_id,
                            defaults={
                                'location': base_loc
                            }
                        )

        if outdated:
            print('error: outdated_version')
            return HttpResponseForbidden(
                json.dumps({'error': 'outdated_version'}),
                content_type='application/json'
            )

        return view_func(request, *args, **kwargs)

    return proxy


@csrf_exempt
@exception_to_500
@transaction.atomic
def create_one_off_code(request):
    phone = normalized_phone(_get_value(request, 'phone'))

    # Todo: check phone?

    try:
        worker = Worker.objects.filter(
            tel_number=phone
        ).select_related(
            'workeruser__user'
        ).get()
        try:
            code = models.update_one_off_code(worker.workeruser.user, phone)
        except WorkerUser.DoesNotExist:
            code = models.create_user_with_one_off_code(phone)
            WorkerUser.objects.create(worker=worker, user=code.user)

        MobileAppWorker.objects.get_or_create(worker_id=worker.pk)
        WorkerRating.objects.get_or_create(worker_id=worker.pk)

    except Worker.DoesNotExist:
        code = models.user_with_one_off_code_by_phone(phone)

    send_sms(
        'GetTask',
        phone,
        'GetTask\nникому не говорите код для входа:\n{}'.format(code.code)
    )
    print(phone)
    print(code.code)

    return HttpResponse('')


@csrf_exempt
@transaction.atomic
def obtain_tokens(request):
    phone = _get_value(request, 'phone')
    key = _get_value(request, 'key')

    print('<{}>'.format(phone))
    print('<{}>'.format(key))
    phone = normalized_phone(phone)
    print('<{}>'.format(phone))

    if phone is None or key is None:
        return HttpResponseForbidden()

    try:
        worker = Worker.objects.get(tel_number=phone)
        one_off_code = models.OneOffCode.objects.filter(
            user__workeruser__worker=worker,
            code=key
        )
        if not one_off_code.exists():
            return HttpResponseForbidden()

    except ObjectDoesNotExist:
        user_phone = models.UserPhone.objects.filter(phone=phone)
        if not user_phone.exists():
            return HttpResponseForbidden()

        user_phone = user_phone.get()

        one_off_code = models.OneOffCode.objects.filter(
            user=user_phone.user,
            code=key
        )
        if not one_off_code.exists():
            return HttpResponseForbidden()

    one_off_code = one_off_code.get()
    refresh = RefreshToken.for_user(one_off_code.user)
    one_off_code.delete()

    return JsonResponse(
        {
            'refresh': str(refresh),
            'access': str(refresh.access_token)
        }
    )


@mobile_api
def upload_registration_info(request):
    session = models.PhotoLoadSession.objects.filter(
        sender=request.user
    )
    user_phone = models.UserPhone.objects.get(user=request.user)
    if session.exists():
        session = session.get()

        rejected_photos = models.PhotoSessionRejectedPhotos.objects.get(session=session)
        photos = models.get_photos(rejected_photos)
        for key, image in request.FILES.items():
            photo = photos.filter(image__endswith=key)
            if photo.exists():
                photo = photo.get()
                photo.set_image(image)
                photo.change_target(session)

        send_tg_message_to_clerks(
            'Пользователь {} обновил фото для валидации в сессии №{}'.format(
                format_phone(user_phone.phone),
                session.pk
            )
        )

    else:
        session = models.PhotoLoadSession.objects.create(
            content_type='worker',
            status='new',
            sender=request.user
        )

        citizenship = _get_value(request, 'citizenship')
        if citizenship != 'Не Россия':
            models.PhotoSessionCitizenship.objects.create(
                session=session,
                citizenship=models.country_by_name(citizenship)
            )

        for key, image in request.FILES.items():
            models.add_photo(session, image)

        send_tg_message_to_clerks(
            'Новый пользователь {} ожидает валидации! Сессия №{}'.format(
                format_phone(user_phone.phone),
                session.pk
            )
        )

    return HttpResponse('')


def _get_mobile_worker(request, with_last_location=False):
    worker_qs = Worker.objects.with_full_name()
    if with_last_location:
        worker_qs.select_related('lastlocation__location')
    return worker_qs.get(workeruser__user=request.user)


def _get_payout_request_info(worker):
    new_payout_request_ts = PayoutRequest.objects.filter(
        paysheet_entry__isnull=True,
        worker=worker,
    ).values(
        'timestamp'
    ).first()
    if new_payout_request_ts is not None:
        return {
            'timestamp': new_payout_request_ts['timestamp'],
            'paysheet_entry': None
        }

    payout_in_progress = PayoutRequest.objects.filter(
        paysheet_entry__paysheet__is_closed=False,
        worker=worker,
    ).values(
        'paysheet_entry',
        'paysheet_entry__amount',
        'timestamp',
    ).first()
    if payout_in_progress is not None:
        dates = TurnoutOperationToPay.objects.annotate(
            has_relevant_operation=Exists(
                Paysheet_v2EntryOperation.objects.filter(
                    operation=OuterRef('operation'),
                    entry=payout_in_progress['paysheet_entry'],
                )
            )
        ).filter(
            has_relevant_operation=True
        ).aggregate(
            Min('turnout__timesheet__sheet_date'),
            Max('turnout__timesheet__sheet_date'),
        )
        return {
            'timestamp': payout_in_progress['timestamp'],
            'paysheet_entry': {
                'amount': payout_in_progress['paysheet_entry__amount'],
                'first_day': dates['turnout__timesheet__sheet_date__min'],
                'last_day': dates['turnout__timesheet__sheet_date__max']
            }
        }

    return None


@mobile_api
def worker_account_info(request):
    try:
        worker_qs = Worker.objects.all(
        ).select_related(
            'worker_account__account'
        ).with_is_online_tomorrow()

        today = timezone.localdate()
        weekday_now = today.isoweekday()
        need_work_tomorrow_check = weekday_now == 5 or weekday_now == 6  # Friday or Saturday

        if need_work_tomorrow_check:
            worker_qs = worker_qs.with_in_weekend_rest_zone()

        worker = worker_qs.get(pk=_get_mobile_worker(request).pk)
        account = worker.worker_account.account
        payout_request = _get_payout_request_info(worker)

        show_tomorrow_work_request_button = not (
            need_work_tomorrow_check and
            worker.in_weekend_rest_zone or
            hasattr(worker, 'banned')
        ) and utils.is_citizenship_migration_ok(
            worker,
            deadline=today + datetime.timedelta(days=1)
        )

        return JsonResponse(
            {
                'status': 'registration_completed',
                'balance': str(-account.turnover_saldo()),
                'full_name': str(worker),
                'payout_request': payout_request,
                'online_tomorrow': worker.is_online_tomorrow,
                'show_tomorrow_work_request_button': show_tomorrow_work_request_button,
            }
        )

    except ObjectDoesNotExist:
        session = models.PhotoLoadSession.objects.filter(sender=request.user)
        if session.exists():
            session = session.get()
            rejected_photos = models.PhotoSessionRejectedPhotos.objects.filter(session=session)
            if rejected_photos.exists():
                rejected_photos = rejected_photos.get()
                photos = list(
                    models.get_photos(rejected_photos).values_list('image', flat=True)
                )
                if len(photos) > 0:
                    return JsonResponse(
                        {
                            'status': 'documents_rejected',
                            'rejected_photos': photos
                        }
                    )
            return JsonResponse({'status': 'documents_in_processing'})
        return JsonResponse({'status': 'new_user'})


def _serialize_selfemployment_data(data):
    return {
        'tax_number': data.tax_number,
        'bank_account': data.bank_account,
        'bank_name': data.bank_name,
        'bank_identification_code': data.bank_identification_code,
        'correspondent_account': data.correspondent_account,
        'cardholder_name': data.cardholder_name,
    }


@mobile_api
def worker_payment_info(request):
    worker = _get_mobile_worker(request)
    try:
        data = worker.selfemployment_data.get(deletion_ts__isnull=True)

    except ObjectDoesNotExist:
        class FakeWSEData:
            def __getattr__(self, name):
                if name == 'cardholder_name':
                    return str(worker)
                else:
                    return ''
        data = FakeWSEData()

    return JsonResponse(_serialize_selfemployment_data(data))


@mobile_api
def update_worker_payment_info(request):
    tax_number = _get_value(request, 'tax_number')
    bank_account = _get_value(request, 'bank_account')
    bank_name = _get_value(request, 'bank_name')
    bank_identification_code = _get_value(request, 'bank_identification_code')
    correspondent_account = _get_value(request, 'correspondent_account')
    cardholder_name = _get_value(request, 'cardholder_name', None)

    worker = _get_mobile_worker(request)
    if cardholder_name is None:
        cardholder_name = worker.full_name

    try:
        data = models.WorkerSelfEmploymentData.objects.get(
            worker=worker,
            deletion_ts__isnull=True
        )

    except ObjectDoesNotExist:
        data = models.WorkerSelfEmploymentData.objects.create(
            worker=worker,
            tax_number=tax_number,
            bank_account=bank_account,
            bank_name=bank_name,
            bank_identification_code=bank_identification_code,
            correspondent_account=correspondent_account,
            cardholder_name=cardholder_name,
        )

        # GT-551: update operations only if selfemployment status changed
        update_not_paid_turnout_payments(worker, request.user)

    else:
        data.tax_number = tax_number
        data.bank_account = bank_account
        data.bank_name = bank_name
        data.bank_identification_code = bank_identification_code
        data.correspondent_account = correspondent_account
        data.cardholder_name = cardholder_name
        data.save()

    return JsonResponse(_serialize_selfemployment_data(data))


@mobile_api
def status_update(request):
    # Nothing to do.
    # Everything is done inside mobile_api annotation.
    return HttpResponse('')


@mobile_api
def update_online_status_by_worker(request):
    online = strtobool(_get_value(request, 'online'))
    try:
        update_online_status(
            is_online=online,
            author=request.user,
            workeruser=request.user.workeruser,
        )
    except NoWorkPermit:
        pass

    return HttpResponse('')


@mobile_api
def request_payout(request):
    worker = Worker.objects.select_for_update(
        of=('self',),
        nowait=True,
    ).with_selfemployment_status(
    ).get(
        workeruser__user=request.user
    )
    if not worker.is_selfemployed:
        return HttpResponseForbidden(
            json.dumps({'error': 'is_not_selfemployed'}),
            content_type='application/json'
        )
    outstanding_requests = PayoutRequest.objects.filter(
        Q(paysheet_entry__isnull=True) |
        Q(paysheet_entry__paysheet__is_closed=False),
        worker=worker,
    )
    if outstanding_requests.exists():
        return HttpResponseForbidden(
            json.dumps({'error': 'payout_already_requested'}),
            content_type='application/json'
        )
    payout_request = PayoutRequest.objects.create(
        worker=worker,
        author=request.user,
    )
    return JsonResponse(
        status=201,
        data={'timestamp': payout_request.timestamp}
    )


# ### Landing

# Todo: move somewhere or refactor
def _do_get_customer_amount(address: str, workers_required: int):
    from the_redhuman_is.async_utils import dadata

    dadata_data = dadata.clean_address(address)

    (lat, lon), in_moscow = tariffs._validate_coordinate_pair(json.dumps(dadata_data), None)

    amount = tariffs.estimate_request_price_for_customer(
        195,
        workers_required,
        None,
        timezone.now().date(),
        timezone.now().time(), # Todo
        [(lat, lon)],
        in_moscow
    )

    return amount


@csrf_exempt
def calc(request):
    params = json.loads(request.body)

    workers_required = int(params.get('moversQty', 1))
    mass = params.get('weight')
    address = params.get('address')

    try:
        amount = _do_get_customer_amount(address, workers_required)
        text = f'{amount} р.'

    except:
        amount = None
        text = 'уточните адрес : ('

    return JsonResponse(
        {
            'status': 'ok',
            'data': {
                'amount': amount,
                'amount_text': text
            }
        }
    )


@csrf_exempt
def create_request_from_landing(request):
    import secrets
    import string
    from the_redhuman_is.tasks import send_email

    params = json.loads(request.body)

    payment_type = 'Неизвестный способ оплаты'
    if params.get('pay') == 'card':
        payment_type = 'Оплата картой'
    elif params.get('pay') == 'legal_entity_invoice':
        payment_type = 'Безналичная оплата от юрлица'

    try:
        user = User.objects.get(username='admin')
        name = params.get('name')
        phone = normalized_phone(params.get('phone'))
        address = params.get('address')
        date = datetime.datetime.strptime(params.get('date'), '%m/%d/%Y').date()
        workers_required=int(params.get('movers_qty', 1))
        shipment_type = params.get('about')

        delivery_request = actions.create_delivery_request(
            user=user,
            customer_id=195,
            location_id=None,
            date=date,
            driver_name=name,
            driver_phones=[phone],
            items=[
                {
                    'interval_begin': params.get('from_time'),
                    'interval_end': params.get('to_time'),
                    'code': ''.join(secrets.choice(string.digits) for i in range(10)),
                    'mass': params.get('weight'),
                    'volume': params.get('volume'),
                    'max_size': None,
                    'place_count': params.get('items_qty'),
                    'shipment_type': shipment_type,
                    'address': address,
                    'workers_required': workers_required,
                }
            ]
        )

        delivery_request.comment = payment_type
        delivery_request.save(user=user)

        # Todo: GT-1065
        _SPECIAL_EMAILS = [
            'ad@gettask.ru',
            'cherepanov.n@gettask.ru',
            'it@gettask.ru',
            'zallexx@yandex.ru',
        ]

        body = 'Имя: {}<br>' \
            'Тел.: {}<br>' \
            'Адрес: {}<br>' \
            'Дата: {}<br>' \
            'Кол-во грузчиков: {}<br>' \
            'Характер груза: {}<br>' \
            'Оплата: {}<br>'.format(
                name,
                phone,
                address,
                date,
                workers_required,
                shipment_type,
                payment_type
            )

        for target in _SPECIAL_EMAILS:
            send_email(
                target,
                f'Новая заявка {delivery_request.pk}',
                body,
                'gt_noreply'
            )

        status = 'ok'
        message = 'Заявка успешно создана. Скоро с вами свяжется наш менеджер!'

    except:
        raise
        status = 'error'
        message = 'Что-то пошло не так, попробуйте позже : ('

    return JsonResponse(
        {
            'status': status,
            'message': message
        }
    )
